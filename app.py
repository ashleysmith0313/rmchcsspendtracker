import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import json
import os
import uuid
import io
from datetime import date, datetime, timedelta
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
)
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

# ══════════════════════════════════════════════════════════════════════════════
# DATA STORE  (no external files needed — JSON lives next to app.py)
# ══════════════════════════════════════════════════════════════════════════════
BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(BASE_DIR, "spend_log.json")

def _ensure_data_file():
    if not os.path.exists(DATA_FILE):
        with open(DATA_FILE, "w") as f:
            json.dump([], f)

# Service lines billed hourly vs daily
HOURLY_SERVICE_LINES = {"Nursing", "Allied Health"}

def load_data() -> pd.DataFrame:
    _ensure_data_file()
    with open(DATA_FILE, "r") as f:
        data = json.load(f)
    if not data:
        return pd.DataFrame(columns=[
            "id", "week_ending", "provider_name", "provider_type",
            "specialty", "service_line", "department",
            "days_worked", "daily_rate",
            "hours_worked", "bill_rate",
            "is_ot", "invoice_number",
            "total_spend", "notes", "logged_at"
        ])
    df = pd.DataFrame(data)
    for col in ["days_worked", "daily_rate", "hours_worked", "bill_rate", "total_spend"]:
        if col not in df.columns:
            df[col] = None
        df[col] = pd.to_numeric(df[col], errors="coerce")
    if "is_ot" not in df.columns:
        df["is_ot"] = False
    if "invoice_number" not in df.columns:
        df["invoice_number"] = ""
    df["is_ot"] = df["is_ot"].fillna(False)
    return df

def save_entry(entry: dict):
    _ensure_data_file()
    with open(DATA_FILE, "r") as f:
        data = json.load(f)
    entry["id"] = str(uuid.uuid4())
    data.append(entry)
    with open(DATA_FILE, "w") as f:
        json.dump(data, f, indent=2)

def delete_entry(entry_id: str):
    _ensure_data_file()
    with open(DATA_FILE, "r") as f:
        data = json.load(f)
    data = [e for e in data if e.get("id") != entry_id]
    with open(DATA_FILE, "w") as f:
        json.dump(data, f, indent=2)

def get_week_ending() -> date:
    today = date.today()
    days_until_saturday = (5 - today.weekday()) % 7
    return today if days_until_saturday == 0 else today + timedelta(days=days_until_saturday)

# ── Requisition & Candidate data store ───────────────────────────────────────
REQ_FILE       = os.path.join(BASE_DIR, "reqs_log.json")
CANDIDATE_FILE = os.path.join(BASE_DIR, "candidates_log.json")

def _ensure(path):
    if not os.path.exists(path):
        with open(path, "w") as f: json.dump([], f)

def _load(path) -> list:
    _ensure(path)
    with open(path, "r") as f: return json.load(f)

def _save_record(path, record):
    _ensure(path)
    data = _load(path)
    record["id"] = str(uuid.uuid4())
    record["created_at"] = datetime.now().isoformat()
    data.append(record)
    with open(path, "w") as f: json.dump(data, f, indent=2)
    return record["id"]

def _update_record(path, record_id, updates):
    _ensure(path)
    data = _load(path)
    for r in data:
        if r.get("id") == record_id:
            r.update(updates)
            r["updated_at"] = datetime.now().isoformat()
    with open(path, "w") as f: json.dump(data, f, indent=2)

def _delete_record(path, record_id):
    _ensure(path)
    data = [r for r in _load(path) if r.get("id") != record_id]
    with open(path, "w") as f: json.dump(data, f, indent=2)

def load_reqs() -> pd.DataFrame:
    data = _load(REQ_FILE)
    if not data:
        return pd.DataFrame(columns=["id","specialty","job_title","discipline","shift","req_type",
                                      "bill_rate","slots_open","req_open_date","status","notes","created_at"])
    return pd.DataFrame(data)

def load_candidates() -> pd.DataFrame:
    data = _load(CANDIDATE_FILE)
    if not data:
        return pd.DataFrame(columns=["id","req_id","candidate_name","source_company","discipline",
                                      "specialty","date_sent","date_clinical_call","date_offered",
                                      "date_accepted","start_date","status","rmchcs_notes",
                                      "cred_company","cred_due_date","cred_status","cred_nm_fingerprint",
                                      "cred_notes","notes","created_at"])
    return pd.DataFrame(data)

# Status option lists
REQ_STATUSES  = ["Open","On Hold","Max Submissions","Filled","Closed"]
CAND_STATUSES = ["Submitted","Clinical Call Scheduled","Clinical Call Complete",
                  "Offered","Accepted","Declined by Candidate","Declined by Client",
                  "Placed","Cancelled","Job Closed"]

# Statuses that are still "active" — candidates in these get auto-closed when req closes
ACTIVE_CAND_STATUSES = {"Submitted","Clinical Call Scheduled","Clinical Call Complete",
                         "Offered","Accepted"}
CRED_STATUSES = ["Pending","Clear","Cancelled","Hold"]
SOURCE_COS    = ["Vista","Springboard","Trustaff","Other IGV Brand","External"]
DISCIPLINES   = ["Nursing","Allied Health","Physician","APP","Locums"]
REQ_TYPES     = ["Backfill","Open Req"]

# ══════════════════════════════════════════════════════════════════════════════
# PIPELINE EXCEL IMPORT HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def _fmt_date(val):
    if val is None: return ""
    if isinstance(val, datetime): return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    return "" if s in ("TBD","ASAP","nan","None","") else s

def _map_cand_status(raw):
    if not raw: return "Submitted"
    r = str(raw).lower().strip()
    if any(x in r for x in ["offered - signed","offered - sign","signed"]): return "Placed"
    if "offered" in r and "pending" in r: return "Offered"
    if "offered" in r: return "Offered"
    if "hired" in r: return "Accepted"
    if "interviewed" in r: return "Clinical Call Complete"
    if "pending clinical call" in r or "clinical call" in r: return "Clinical Call Scheduled"
    if "completed clinical" in r: return "Clinical Call Complete"
    if "accepted another offer" in r or "declined - rto" in r: return "Declined by Candidate"
    if "did not call" in r or "would not recommend" in r or "no total hips" in r or ", pass" in r: return "Declined by Client"
    if "cancel" in r: return "Cancelled"
    if "sent to mgr" in r: return "Submitted"
    return "Submitted"

def _map_req_status(raw):
    if not raw: return "Open"
    r = str(raw).lower().strip()
    if "closed" in r: return "Closed"
    if "filled" in r: return "Filled"
    if "max sub" in r: return "Max Submissions"
    if "hold" in r: return "On Hold"
    if "pending approval" in r: return "On Hold"
    if "open" in r: return "Open"
    return "Open"

def _map_cred_status(raw):
    if not raw: return ""
    r = str(raw).lower()
    if "cancel" in r: return "Cancelled"
    if "clear" in r: return "Clear"
    if "hold" in r: return "Hold"
    return "Pending"

def _infer_discipline(specialty):
    if not specialty: return "Nursing"
    s = specialty.lower()
    if any(x in s for x in ["surgery","ortho","surgeon","physician","locum","medicine","md "]): return "Physician"
    if any(x in s for x in ["tech","rad","nuc","rrt","crt","cst","surg tech","ultrasound","respiratory","echo","phlebotomy"]): return "Allied Health"
    return "Nursing"

def _clean_name(name):
    return str(name).replace("*SB","").replace("*sb","").strip() if name else ""

def _infer_source(name):
    return "Springboard" if name and "*sb" in str(name).lower() else "Vista"

def parse_pipeline_excel(file_obj) -> dict:
    """Parse the RMCHCS candidates/jobs Excel. Returns {reqs, candidates, warnings}."""
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    reqs, candidates, warnings = [], [], []

    # ── Requisitions ──────────────────────────────────────────────────────
    if "Open Needs and Backfill Needs" in wb.sheetnames:
        ws = wb["Open Needs and Backfill Needs"]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            if not any(v is not None for v in row): continue
            # Left table: Allied/Nursing (cols 0-7)
            if row[0] and str(row[0]).strip() not in ("","Specialty"):
                disc_raw = str(row[0]).strip()
                disc = "Allied Health" if disc_raw.lower()=="allied" else "Nursing" if disc_raw.lower()=="nursing" else disc_raw
                try: br = float(str(row[2]).replace("$","").replace("/hr","").strip()) if row[2] else 0
                except: br = 0
                req_type = "Backfill" if "backfill" in str(row[6] or "").lower() else "Open Req"
                slots = 1
                try: slots = int(row[5]) if row[5] is not None else 1
                except: pass
                reqs.append({
                    "specialty": str(row[1]).strip() if row[1] else "",
                    "job_title": str(row[1]).strip() if row[1] else "",
                    "discipline": disc, "shift": str(row[3]).strip() if row[3] else "",
                    "req_type": req_type, "bill_rate": br, "slots_open": slots,
                    "req_open_date": "", "status": _map_req_status(row[4]),
                    "notes": str(row[7]).strip() if row[7] else ""
                })
            # Right table: Locums (cols 9-14)
            if row[9] and str(row[9]).strip() not in ("","Specialty"):
                try: br = float(str(row[11]).replace("$","").replace("/hr","").strip()) if row[11] else 0
                except: br = 0
                reqs.append({
                    "specialty": str(row[10]).strip() if row[10] else "",
                    "job_title": str(row[10]).strip() if row[10] else "",
                    "discipline": "Physician", "shift": "",
                    "req_type": "Open Req", "bill_rate": br, "slots_open": 1,
                    "req_open_date": "", "status": _map_req_status(row[13]),
                    "notes": str(row[14]).strip() if row[14] else ""
                })
    else:
        warnings.append("Sheet 'Open Needs and Backfill Needs' not found — reqs skipped.")

    # ── Credentialing lookup ──────────────────────────────────────────────
    cred_lookup = {}
    if "Credentialing" in wb.sheetnames:
        ws_c = wb["Credentialing"]
        for row in list(ws_c.iter_rows(values_only=True))[1:]:
            if not row[3]: continue
            raw_name = str(row[3]).strip()
            base = raw_name.replace("- CANCEL","").replace("(Ext flip)","").replace("- pending","").replace("- Cancel","").strip().lower()
            cred_lookup[base] = {
                "cred_company":        str(row[0]).strip() if row[0] else "",
                "cred_due_date":       _fmt_date(row[2]),
                "cred_status":         _map_cred_status(str(row[5]) if row[5] else ""),
                "cred_nm_fingerprint": bool(row[7]) if row[7] else False,
                "cred_notes":          str(row[5]).strip() if row[5] else ""
            }

    # ── Candidates ────────────────────────────────────────────────────────
    def _parse_cand_sheet(ws, disc_override=None):
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            if not row[0]: continue
            raw_name = str(row[0]).strip()
            if raw_name.lower() in ("candidate","name",""): continue
            name    = _clean_name(raw_name)
            source  = _infer_source(raw_name)
            spec    = str(row[1]).strip() if row[1] else ""
            disc    = disc_override or _infer_discipline(spec)
            d_sent  = _fmt_date(row[2])
            if disc_override == "Physician":
                notes_raw = ""
                rmchcs    = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                start_date= ""
            else:
                notes_raw = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                rmchcs    = str(row[5]).strip() if len(row) > 5 and row[5] else ""
                start_date= _fmt_date(row[3]) if len(row) > 3 and row[3] else ""
            status = _map_cand_status(notes_raw or rmchcs)
            cred   = cred_lookup.get(name.lower(), {})
            candidates.append({
                "req_id": None,
                "candidate_name": name, "source_company": source,
                "discipline": disc, "specialty": spec, "status": status,
                "date_sent": d_sent, "date_clinical_call": "",
                "date_offered": "", "date_accepted": "", "start_date": start_date,
                "notes": notes_raw, "rmchcs_notes": rmchcs,
                "cred_company":        cred.get("cred_company",""),
                "cred_due_date":       cred.get("cred_due_date",""),
                "cred_status":         cred.get("cred_status",""),
                "cred_nm_fingerprint": cred.get("cred_nm_fingerprint",False),
                "cred_notes":          cred.get("cred_notes","")
            })

    if "Travel Nurse and Allied" in wb.sheetnames:
        _parse_cand_sheet(wb["Travel Nurse and Allied"])
    else:
        warnings.append("Sheet 'Travel Nurse and Allied' not found.")
    if "Physicians - Locums" in wb.sheetnames:
        _parse_cand_sheet(wb["Physicians - Locums"], disc_override="Physician")
    else:
        warnings.append("Sheet 'Physicians - Locums' not found.")

    return {"reqs": reqs, "candidates": candidates, "warnings": warnings}

# ══════════════════════════════════════════════════════════════════════════════
# INVOICE PARSER HELPERS
# ══════════════════════════════════════════════════════════════════════════════
_ALLIED_KW  = ["ultrasound","radiology","lab","tech","therapy","respiratory","imaging",
               "surgical","echo","cardio","vascular","mri","ct ","nuclear","phlebotomy",
               "x-ray","xray","mammography","ekg","eeg","sterile","supply","pharmacy"]
_NURSING_KW = [" rn","lpn","nurse","nursing","cna","med/surg","icu","er ","ed ","l&d",
               "labor","oncology","telemetry","pacu","or ","post op","step down","float"]

def _invoice_service_line(placement: str) -> str:
    p = placement.lower()
    for kw in _ALLIED_KW:
        if kw in p:
            return "Allied Health"
    for kw in _NURSING_KW:
        if kw in p:
            return "Nursing"
    return "Allied Health"

def _invoice_specialty(placement: str) -> str:
    parts = placement.split("|")
    if len(parts) >= 3:
        return parts[-1].strip()
    return placement.strip()

def _invoice_provider_type(placement: str) -> str:
    p = placement.lower()
    if "tech" in p:       return "Tech"
    if " rn" in p:        return "RN"
    if "lpn" in p:        return "LPN"
    if "cna" in p:        return "CNA"
    if "therapist" in p:  return "Therapist"
    if "nurse" in p:      return "RN"
    return "Other"

def parse_invoice_xlsx(file_obj) -> tuple:
    """Parse a Vista consolidated invoice Excel file.
    Returns (parsed_rows: list[dict], warnings: list[str], skipped: int)"""
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], ["File appears empty."], 0

    headers = [str(h).strip().lower() if h else "" for h in all_rows[0]]

    # Flexible column detection
    def col(name_options):
        for name in name_options:
            for i, h in enumerate(headers):
                if name in h:
                    return i
        return None

    idx_nurse    = col(["nurse name", "provider name", "worker"])
    idx_we       = col(["we date", "week end", "week ending"])
    idx_placement= col(["placement"])
    idx_rev_cat  = col(["revenue category", "rev cat", "category"])
    idx_qty      = col(["quantity", "qty", "hours"])
    idx_rate     = col(["unit cost", "bill rate", "rate"])
    idx_amount   = col(["line amount", "amount"])
    idx_invoice  = col(["invoice number", "invoice #", "inv number"])

    missing_cols = []
    for name, idx in [("Nurse Name", idx_nurse), ("WE Date", idx_we),
                      ("Revenue Category", idx_rev_cat), ("Quantity", idx_qty),
                      ("Unit Cost", idx_rate), ("Line Amount", idx_amount)]:
        if idx is None:
            missing_cols.append(name)
    if missing_cols:
        return [], [f"Could not find required columns: {', '.join(missing_cols)}"], 0

    parsed = []
    skipped = 0
    warnings = []

    for row in all_rows[1:]:
        if not any(v is not None for v in row):
            continue
        nurse_name  = row[idx_nurse]    if idx_nurse    is not None else None
        we_date     = row[idx_we]       if idx_we       is not None else None
        placement   = row[idx_placement]if idx_placement is not None else ""
        rev_cat     = row[idx_rev_cat]  if idx_rev_cat  is not None else ""
        quantity    = row[idx_qty]      if idx_qty      is not None else None
        unit_cost   = row[idx_rate]     if idx_rate     is not None else None
        line_amount = row[idx_amount]   if idx_amount   is not None else None
        invoice_num = row[idx_invoice]  if idx_invoice  is not None else ""

        # Skip summary/total rows
        if nurse_name is None or str(nurse_name).strip() == "":
            skipped += 1
            continue
        if not isinstance(we_date, datetime):
            skipped += 1
            continue

        try:
            qty_f    = float(quantity)   if quantity   is not None else 0.0
            rate_f   = float(unit_cost)  if unit_cost  is not None else 0.0
            amount_f = round(float(line_amount), 2) if line_amount is not None else round(qty_f * rate_f, 2)
        except (TypeError, ValueError):
            skipped += 1
            warnings.append(f"Skipped row for {nurse_name} — could not parse numeric values.")
            continue

        placement_str = str(placement) if placement else str(nurse_name)
        is_ot = "over time" in str(rev_cat).lower() or "overtime" in str(rev_cat).lower()

        parsed.append({
            "week_ending":    we_date.strftime("%Y-%m-%d"),
            "provider_name":  str(nurse_name).strip(),
            "provider_type":  _invoice_provider_type(placement_str),
            "specialty":      _invoice_specialty(placement_str),
            "service_line":   _invoice_service_line(placement_str),
            "department":     "",
            "days_worked":    None,
            "daily_rate":     None,
            "hours_worked":   qty_f,
            "bill_rate":      rate_f,
            "is_ot":          is_ot,
            "invoice_number": str(invoice_num).strip() if invoice_num else "",
            "total_spend":    amount_f,
            "notes":          f"Invoice {invoice_num} | {str(rev_cat).strip()}" if invoice_num else str(rev_cat).strip(),
            "logged_at":      datetime.now().isoformat()
        })

    return parsed, warnings, skipped

# ══════════════════════════════════════════════════════════════════════════════
# PDF GENERATOR
# ══════════════════════════════════════════════════════════════════════════════
NAVY      = colors.HexColor("#0f1724")
BLUE      = colors.HexColor("#1a3a5c")
SLATE     = colors.HexColor("#64748b")
LIGHT     = colors.HexColor("#f1f5f9")
WHITE     = colors.white
BORDER    = colors.HexColor("#e2e8f0")

def generate_pdf_report(df, week_ending, title, prepared_by="Ingenovis ITO",
                         include_detail=True, include_notes=True) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
                             rightMargin=0.65*inch, leftMargin=0.65*inch,
                             topMargin=0.65*inch, bottomMargin=0.65*inch)

    header_style   = ParagraphStyle("h",  fontName="Helvetica-Bold", fontSize=18, textColor=WHITE,   leading=22, alignment=TA_LEFT)
    sub_style      = ParagraphStyle("s",  fontName="Helvetica",      fontSize=9,  textColor=colors.HexColor("#94a3b8"), leading=13)
    section_style  = ParagraphStyle("sc", fontName="Helvetica-Bold", fontSize=11, textColor=NAVY,    leading=14, spaceBefore=14, spaceAfter=6)
    kpi_lbl_style  = ParagraphStyle("kl", fontName="Helvetica",      fontSize=7,  textColor=SLATE,   leading=10, alignment=TA_CENTER)
    kpi_val_style  = ParagraphStyle("kv", fontName="Helvetica-Bold", fontSize=16, textColor=NAVY,    leading=20, alignment=TA_CENTER)
    small_style    = ParagraphStyle("sm", fontName="Helvetica",      fontSize=8,  textColor=SLATE,   leading=11)
    right_style    = ParagraphStyle("r",  fontName="Helvetica",      fontSize=8,  textColor=SLATE,   leading=11, alignment=TA_RIGHT)

    story = []

    # Header banner
    hdr = Table([[
        Paragraph("<b>RMCHCS Weekly Spend Report</b>", header_style),
        Paragraph(f"Week Ending: {week_ending}<br/>Generated: {datetime.now().strftime('%B %d, %Y')}<br/>Prepared by: {prepared_by}", sub_style)
    ]], colWidths=[4.2*inch, 2.8*inch])
    hdr.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), NAVY),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING",(0,0), (0,-1),  18),
        ("RIGHTPADDING",(-1,0),(-1,-1),16),
        ("TOPPADDING", (0,0), (-1,-1), 16),
        ("BOTTOMPADDING",(0,0),(-1,-1),16),
    ]))
    story += [hdr, Spacer(1, 0.2*inch)]

    # KPI row
    total_spend     = df["total_spend"].sum()
    total_providers = df["provider_name"].nunique()
    total_specs     = df["specialty"].nunique()

    has_hourly = df["hours_worked"].notna().any() and (df["hours_worked"].fillna(0) > 0).any()
    has_daily  = df["days_worked"].notna().any()  and (df["days_worked"].fillna(0)  > 0).any()
    if has_hourly and not has_daily:
        kpi4_label = "TOTAL HOURS"
        kpi4_val   = f"{df['hours_worked'].fillna(0).sum():.1f} hrs"
    elif has_daily and not has_hourly:
        kpi4_label = "TOTAL DAYS"
        kpi4_val   = f"{df['days_worked'].fillna(0).sum():.1f} days"
    else:
        kpi4_label = "TOTAL HOURS"
        tot_hrs_v  = df["hours_worked"].fillna(0).sum()
        tot_days_v = df["days_worked"].fillna(0).sum()
        kpi4_val   = f"{tot_hrs_v:.1f} hrs / {tot_days_v:.1f} days"

    kpi_items = [("TOTAL SPEND", f"${total_spend:,.2f}"),
                 ("PROVIDERS",   str(total_providers)),
                 ("SPECIALTIES", str(total_specs)),
                 (kpi4_label,    kpi4_val)]
    kpi_cells = [[Paragraph(l, kpi_lbl_style) for l,_ in kpi_items],
                 [Paragraph(v, kpi_val_style) for _,v in kpi_items]]
    kpi_tbl = Table(kpi_cells, colWidths=[1.74*inch]*4)
    kpi_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(-1,-1), LIGHT),
        ("BOX",        (0,0),(-1,-1), 1, BORDER),
        ("LINEAFTER",  (0,0),(2,-1),  0.5, BORDER),
        ("TOPPADDING", (0,0),(-1,-1), 10),
        ("BOTTOMPADDING",(0,0),(-1,-1),10),
        ("VALIGN",     (0,0),(-1,-1), "MIDDLE"),
    ]))
    story += [kpi_tbl, Spacer(1, 0.18*inch)]

    # Specialty summary
    story.append(Paragraph("Spend by Specialty", section_style))
    spec_hrs_map  = df.groupby("specialty")["hours_worked"].sum()
    spec_days_map = df.groupby("specialty")["days_worked"].sum()
    spec = df.groupby("specialty").agg(
        Providers=("provider_name","nunique"),
        Spend=("total_spend","sum")
    ).reset_index().sort_values("Spend", ascending=False)
    spec["Hours"] = spec["specialty"].map(spec_hrs_map).fillna(0)
    spec["Days"]  = spec["specialty"].map(spec_days_map).fillna(0)

    if has_hourly and not has_daily:
        qty_col_hdr = "Hours Worked"
    elif has_daily and not has_hourly:
        qty_col_hdr = "Days Worked"
    else:
        qty_col_hdr = "Hrs / Days"

    spec_rows = [["Specialty","Providers", qty_col_hdr, "Total Spend","% of Week"]]
    for _, row in spec.iterrows():
        pct = (row["Spend"]/total_spend*100) if total_spend else 0
        if has_hourly and not has_daily:
            qty_str = f"{row['Hours']:.1f} hrs"
        elif has_daily and not has_hourly:
            qty_str = f"{row['Days']:.1f} days"
        else:
            qty_str = f"{row['Hours']:.1f}h / {row['Days']:.1f}d"
        spec_rows.append([row["specialty"], str(int(row["Providers"])),
                          qty_str, f"${row['Spend']:,.2f}", f"{pct:.1f}%"])
    tot_hrs_s  = df["hours_worked"].fillna(0).sum()
    tot_days_s = df["days_worked"].fillna(0).sum()
    if has_hourly and not has_daily:
        tot_qty_str = f"{tot_hrs_s:.1f} hrs"
    elif has_daily and not has_hourly:
        tot_qty_str = f"{tot_days_s:.1f} days"
    else:
        tot_qty_str = f"{tot_hrs_s:.1f}h / {tot_days_s:.1f}d"
    spec_rows.append(["TOTAL", str(total_providers), tot_qty_str, f"${total_spend:,.2f}", "100%"])

    st_tbl = Table(spec_rows, colWidths=[2.4*inch,1.0*inch,1.0*inch,1.3*inch,0.95*inch])
    st_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),BLUE), ("TEXTCOLOR",(0,0),(-1,0),WHITE),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"), ("FONTSIZE",(0,0),(-1,0),8),
        ("FONTNAME",(0,1),(-1,-1),"Helvetica"),     ("FONTSIZE",(0,1),(-1,-1),8),
        ("TEXTCOLOR",(0,1),(-1,-1),NAVY),
        ("ROWBACKGROUNDS",(0,1),(-1,-2),[WHITE,LIGHT]),
        ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#dbeafe")),
        ("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
        ("ALIGN",(1,0),(-1,-1),"CENTER"), ("ALIGN",(3,0),(3,-1),"RIGHT"),
        ("TOPPADDING",(0,0),(-1,-1),6), ("BOTTOMPADDING",(0,0),(-1,-1),6),
        ("LEFTPADDING",(0,0),(0,-1),10), ("GRID",(0,0),(-1,-1),0.4,BORDER),
    ]))
    story += [st_tbl, Spacer(1, 0.15*inch)]

    # Service line summary
    story.append(Paragraph("Spend by Service Line", section_style))
    sl = df.groupby("service_line").agg(
        Providers=("provider_name","nunique"),
        Spend=("total_spend","sum")
    ).reset_index().sort_values("Spend", ascending=False)

    sl_rows = [["Service Line","Providers","Total Spend","% of Week"]]
    for _, row in sl.iterrows():
        pct = (row["Spend"]/total_spend*100) if total_spend else 0
        sl_rows.append([row["service_line"], str(int(row["Providers"])),
                        f"${row['Spend']:,.2f}", f"{pct:.1f}%"])
    sl_tbl = Table(sl_rows, colWidths=[2.4*inch,1.1*inch,1.5*inch,1.1*inch])
    sl_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),BLUE), ("TEXTCOLOR",(0,0),(-1,0),WHITE),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"), ("FONTSIZE",(0,0),(-1,0),8),
        ("FONTNAME",(0,1),(-1,-1),"Helvetica"),     ("FONTSIZE",(0,1),(-1,-1),8),
        ("TEXTCOLOR",(0,1),(-1,-1),NAVY),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[WHITE,LIGHT]),
        ("ALIGN",(1,0),(-1,-1),"CENTER"), ("ALIGN",(2,0),(2,-1),"RIGHT"),
        ("TOPPADDING",(0,0),(-1,-1),6), ("BOTTOMPADDING",(0,0),(-1,-1),6),
        ("LEFTPADDING",(0,0),(0,-1),10), ("GRID",(0,0),(-1,-1),0.4,BORDER),
    ]))
    story.append(sl_tbl)

    # Provider detail
    if include_detail:
        story += [Spacer(1, 0.1*inch), Paragraph("Provider Detail", section_style)]
        hdr_row = [["Provider","Type","Specialty","Svc Line","Hrs / Days","Rate","Total"]]
        if include_notes:
            hdr_row[0].append("Notes")
        det_rows = []
        for _, row in df.sort_values(["specialty","provider_name"]).iterrows():
            sl = row.get("service_line","")
            if sl in HOURLY_SERVICE_LINES:
                qty  = f"{row['hours_worked']:.2f} hrs" if pd.notna(row.get("hours_worked")) else ""
                rate = f"${row['bill_rate']:,.2f}/hr"   if pd.notna(row.get("bill_rate"))    else ""
            else:
                qty  = f"{row['days_worked']:.1f} days" if pd.notna(row.get("days_worked")) else ""
                rate = f"${row['daily_rate']:,.2f}/day"  if pd.notna(row.get("daily_rate"))  else ""
            ot_flag = " (OT)" if row.get("is_ot") else ""
            r = [row["provider_name"] + ot_flag, row.get("provider_type",""),
                 row["specialty"], sl, qty, rate, f"${row['total_spend']:,.2f}"]
            if include_notes:
                r.append(str(row.get("notes","") or ""))
            det_rows.append(r)
        # Wrap notes in Paragraph so long strings word-wrap instead of overflow
        if include_notes:
            notes_style = ParagraphStyle("ns", fontName="Helvetica", fontSize=7, textColor=SLATE, leading=9)
            for r in det_rows:
                r[-1] = Paragraph(str(r[-1]), notes_style)
            # Col widths: Provider, Type, Specialty, SvcLine, Hrs/Days, Rate, Total, Notes
            # Total usable width = 7.0 inch (letter - margins)
            cw = [1.3*inch,0.55*inch,0.95*inch,0.65*inch,0.7*inch,0.82*inch,0.78*inch,1.25*inch]
        else:
            cw = [1.8*inch,0.75*inch,1.35*inch,0.85*inch,0.7*inch,0.9*inch,0.85*inch]
        det_tbl = Table(hdr_row+det_rows, colWidths=cw, repeatRows=1)
        det_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),BLUE), ("TEXTCOLOR",(0,0),(-1,0),WHITE),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"), ("FONTSIZE",(0,0),(-1,0),7.5),
            ("FONTNAME",(0,1),(-1,-1),"Helvetica"),     ("FONTSIZE",(0,1),(-1,-1),7.5),
            ("TEXTCOLOR",(0,1),(-1,-1),NAVY),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[WHITE,LIGHT]),
            ("ALIGN",(4,0),(5,-1),"CENTER"),
            ("ALIGN",(6,0),(6,-1),"RIGHT"),
            ("RIGHTPADDING",(6,0),(6,-1),8),
            ("VALIGN",(0,0),(-1,-1),"TOP"),
            ("TOPPADDING",(0,0),(-1,-1),5), ("BOTTOMPADDING",(0,0),(-1,-1),5),
            ("LEFTPADDING",(0,0),(0,-1),8), ("GRID",(0,0),(-1,-1),0.3,BORDER),
        ]))
        story.append(det_tbl)

    # Footer
    story += [Spacer(1,0.25*inch), HRFlowable(width="100%",thickness=0.5,color=BORDER), Spacer(1,0.08*inch)]
    ft = Table([[Paragraph(f"Confidential | {prepared_by}", small_style),
                 Paragraph(f"Report generated {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", right_style)]],
               colWidths=[3.5*inch, 3.5*inch])
    ft.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),
                             ("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),0)]))
    story.append(ft)

    doc.build(story)
    return buf.getvalue()


def generate_cumulative_report(df, date_from, date_to, title, prepared_by="Ingenovis ITO",
                                include_detail=True, include_notes=True) -> bytes:
    """Generate a cumulative spend report across all weeks in the date range."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
                             rightMargin=0.65*inch, leftMargin=0.65*inch,
                             topMargin=0.65*inch, bottomMargin=0.65*inch)

    header_style  = ParagraphStyle("h2", fontName="Helvetica-Bold", fontSize=18, textColor=WHITE,  leading=22, alignment=TA_LEFT)
    sub_style     = ParagraphStyle("s2", fontName="Helvetica",      fontSize=9,  textColor=colors.HexColor("#94a3b8"), leading=13)
    section_style = ParagraphStyle("sc2",fontName="Helvetica-Bold", fontSize=11, textColor=NAVY,   leading=14, spaceBefore=14, spaceAfter=6)
    kpi_lbl_style = ParagraphStyle("kl2",fontName="Helvetica",      fontSize=7,  textColor=SLATE,  leading=10, alignment=TA_CENTER)
    kpi_val_style = ParagraphStyle("kv2",fontName="Helvetica-Bold", fontSize=16, textColor=NAVY,   leading=20, alignment=TA_CENTER)
    tbl_hdr_style = ParagraphStyle("th2",fontName="Helvetica",      fontSize=8,  textColor=SLATE,  leading=11)
    small_style   = ParagraphStyle("sm2",fontName="Helvetica",      fontSize=8,  textColor=SLATE,  leading=11)
    right_style   = ParagraphStyle("r2", fontName="Helvetica",      fontSize=8,  textColor=SLATE,  leading=11, alignment=TA_RIGHT)

    story = []

    # Header banner
    date_range_str = f"{date_from} through {date_to}"
    hdr = Table([[
        Paragraph(f"<b>{title}</b>", header_style),
        Paragraph(f"Period: {date_range_str}<br/>Generated: {datetime.now().strftime('%B %d, %Y')}<br/>Prepared by: {prepared_by}", sub_style)
    ]], colWidths=[3.8*inch, 3.2*inch])
    hdr.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), NAVY),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING",(0,0), (0,-1),  18),
        ("RIGHTPADDING",(-1,0),(-1,-1),16),
        ("TOPPADDING", (0,0), (-1,-1), 16),
        ("BOTTOMPADDING",(0,0),(-1,-1),16),
    ]))
    story += [hdr, Spacer(1, 0.2*inch)]

    # KPI row
    total_spend     = df["total_spend"].sum()
    total_providers = df["provider_name"].nunique()
    total_specs     = df["specialty"].nunique()
    total_weeks     = df["week_ending"].nunique()

    has_hourly = df["hours_worked"].notna().any() and (df["hours_worked"].fillna(0) > 0).any()
    has_daily  = df["days_worked"].notna().any()  and (df["days_worked"].fillna(0)  > 0).any()

    kpi_items = [
        ("TOTAL SPEND",   f"${total_spend:,.2f}"),
        ("PROVIDERS",     str(total_providers)),
        ("WEEKS TRACKED", str(total_weeks)),
        ("SPECIALTIES",   str(total_specs)),
    ]
    kpi_cells = [[Paragraph(l, kpi_lbl_style) for l,_ in kpi_items],
                 [Paragraph(v, kpi_val_style) for _,v in kpi_items]]
    kpi_tbl = Table(kpi_cells, colWidths=[1.74*inch]*4)
    kpi_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(-1,-1), LIGHT),
        ("BOX",        (0,0),(-1,-1), 1, BORDER),
        ("LINEAFTER",  (0,0),(2,-1),  0.5, BORDER),
        ("TOPPADDING", (0,0),(-1,-1), 10),
        ("BOTTOMPADDING",(0,0),(-1,-1),10),
        ("VALIGN",     (0,0),(-1,-1), "MIDDLE"),
    ]))
    story += [kpi_tbl, Spacer(1, 0.18*inch)]

    # Weekly spend summary table
    story.append(Paragraph("Weekly Spend Summary", section_style))
    weekly = df.groupby("week_ending").agg(
        Providers=("provider_name","nunique"),
        Spend=("total_spend","sum")
    ).reset_index().sort_values("week_ending")
    if has_hourly:
        weekly_hrs  = df.groupby("week_ending")["hours_worked"].sum()
        weekly["Hours"] = weekly["week_ending"].map(weekly_hrs).fillna(0)
    if has_daily:
        weekly_days = df.groupby("week_ending")["days_worked"].sum()
        weekly["Days"] = weekly["week_ending"].map(weekly_days).fillna(0)

    if has_hourly and has_daily:
        wk_rows = [["Week Ending","Providers","Hours","Days","Spend","Cumulative"]]
    elif has_hourly:
        wk_rows = [["Week Ending","Providers","Hours","Spend","Cumulative"]]
    else:
        wk_rows = [["Week Ending","Providers","Days","Spend","Cumulative"]]

    running = 0
    for _, row in weekly.iterrows():
        running += row["Spend"]
        if has_hourly and has_daily:
            wk_rows.append([row["week_ending"],
                            str(int(row["Providers"])),
                            f"{row.get('Hours',0):.1f}",
                            f"{row.get('Days',0):.1f}",
                            f"${row['Spend']:,.2f}",
                            f"${running:,.2f}"])
        elif has_hourly:
            wk_rows.append([row["week_ending"],
                            str(int(row["Providers"])),
                            f"{row.get('Hours',0):.1f} hrs",
                            f"${row['Spend']:,.2f}",
                            f"${running:,.2f}"])
        else:
            wk_rows.append([row["week_ending"],
                            str(int(row["Providers"])),
                            f"{row.get('Days',0):.1f} days",
                            f"${row['Spend']:,.2f}",
                            f"${running:,.2f}"])
    wk_rows.append(["TOTAL", str(total_providers), "", f"${total_spend:,.2f}", ""] if not (has_hourly and has_daily)
                   else ["TOTAL", str(total_providers), "", "", f"${total_spend:,.2f}", ""])

    if has_hourly and has_daily:
        wk_cw = [1.1*inch,0.85*inch,0.7*inch,0.7*inch,1.1*inch,1.2*inch]
    else:
        wk_cw = [1.2*inch,0.9*inch,1.0*inch,1.2*inch,1.35*inch]

    wk_tbl = Table(wk_rows, colWidths=wk_cw)
    wk_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),BLUE), ("TEXTCOLOR",(0,0),(-1,0),WHITE),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"), ("FONTSIZE",(0,0),(-1,0),8),
        ("FONTNAME",(0,1),(-1,-1),"Helvetica"),     ("FONTSIZE",(0,1),(-1,-1),8),
        ("TEXTCOLOR",(0,1),(-1,-1),NAVY),
        ("ROWBACKGROUNDS",(0,1),(-1,-2),[WHITE,LIGHT]),
        ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#dbeafe")),
        ("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
        ("ALIGN",(1,0),(-1,-1),"CENTER"),
        ("ALIGN",(-2,0),(-1,-1),"RIGHT"),
        ("TOPPADDING",(0,0),(-1,-1),6), ("BOTTOMPADDING",(0,0),(-1,-1),6),
        ("LEFTPADDING",(0,0),(0,-1),10), ("GRID",(0,0),(-1,-1),0.4,BORDER),
    ]))
    story += [wk_tbl, Spacer(1, 0.15*inch)]

    # Specialty summary
    story.append(Paragraph("Spend by Specialty", section_style))
    spec_hrs_map  = df.groupby("specialty")["hours_worked"].sum()
    spec_days_map = df.groupby("specialty")["days_worked"].sum()
    spec = df.groupby("specialty").agg(
        Providers=("provider_name","nunique"),
        Weeks=("week_ending","nunique"),
        Spend=("total_spend","sum")
    ).reset_index().sort_values("Spend", ascending=False)
    spec["Hours"] = spec["specialty"].map(spec_hrs_map).fillna(0)
    spec["Days"]  = spec["specialty"].map(spec_days_map).fillna(0)

    if has_hourly and not has_daily:
        sp_rows = [["Specialty","Providers","Weeks","Hours","Total Spend","% of Total"]]
    elif has_daily and not has_hourly:
        sp_rows = [["Specialty","Providers","Weeks","Days","Total Spend","% of Total"]]
    else:
        sp_rows = [["Specialty","Providers","Weeks","Hrs / Days","Total Spend","% of Total"]]

    for _, row in spec.iterrows():
        pct = (row["Spend"]/total_spend*100) if total_spend else 0
        if has_hourly and not has_daily:
            qty_str = f"{row['Hours']:.1f} hrs"
        elif has_daily and not has_hourly:
            qty_str = f"{row['Days']:.1f} days"
        else:
            qty_str = f"{row['Hours']:.1f}h/{row['Days']:.1f}d"
        sp_rows.append([row["specialty"], str(int(row["Providers"])),
                        str(int(row["Weeks"])), qty_str,
                        f"${row['Spend']:,.2f}", f"{pct:.1f}%"])
    tot_hrs_c  = df["hours_worked"].fillna(0).sum()
    tot_days_c = df["days_worked"].fillna(0).sum()
    if has_hourly and not has_daily:   tot_qty_c = f"{tot_hrs_c:.1f} hrs"
    elif has_daily and not has_hourly: tot_qty_c = f"{tot_days_c:.1f} days"
    else:                              tot_qty_c = f"{tot_hrs_c:.1f}h/{tot_days_c:.1f}d"
    sp_rows.append(["TOTAL", str(total_providers), str(total_weeks), tot_qty_c, f"${total_spend:,.2f}", "100%"])

    sp_tbl = Table(sp_rows, colWidths=[1.9*inch,0.8*inch,0.65*inch,0.9*inch,1.1*inch,0.8*inch])
    sp_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),BLUE), ("TEXTCOLOR",(0,0),(-1,0),WHITE),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"), ("FONTSIZE",(0,0),(-1,0),8),
        ("FONTNAME",(0,1),(-1,-1),"Helvetica"),     ("FONTSIZE",(0,1),(-1,-1),8),
        ("TEXTCOLOR",(0,1),(-1,-1),NAVY),
        ("ROWBACKGROUNDS",(0,1),(-1,-2),[WHITE,LIGHT]),
        ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#dbeafe")),
        ("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
        ("ALIGN",(1,0),(-1,-1),"CENTER"), ("ALIGN",(4,0),(4,-1),"RIGHT"),
        ("TOPPADDING",(0,0),(-1,-1),6), ("BOTTOMPADDING",(0,0),(-1,-1),6),
        ("LEFTPADDING",(0,0),(0,-1),10), ("GRID",(0,0),(-1,-1),0.4,BORDER),
    ]))
    story += [sp_tbl, Spacer(1, 0.15*inch)]

    # Service line summary
    story.append(Paragraph("Spend by Service Line", section_style))
    sl_sum = df.groupby("service_line").agg(
        Providers=("provider_name","nunique"),
        Weeks=("week_ending","nunique"),
        Spend=("total_spend","sum")
    ).reset_index().sort_values("Spend", ascending=False)

    sl_rows = [["Service Line","Providers","Weeks","Total Spend","% of Total"]]
    for _, row in sl_sum.iterrows():
        pct = (row["Spend"]/total_spend*100) if total_spend else 0
        sl_rows.append([row["service_line"], str(int(row["Providers"])),
                        str(int(row["Weeks"])), f"${row['Spend']:,.2f}", f"{pct:.1f}%"])
    sl_tbl = Table(sl_rows, colWidths=[1.9*inch,0.9*inch,0.75*inch,1.5*inch,1.1*inch])
    sl_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),BLUE), ("TEXTCOLOR",(0,0),(-1,0),WHITE),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"), ("FONTSIZE",(0,0),(-1,0),8),
        ("FONTNAME",(0,1),(-1,-1),"Helvetica"),     ("FONTSIZE",(0,1),(-1,-1),8),
        ("TEXTCOLOR",(0,1),(-1,-1),NAVY),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[WHITE,LIGHT]),
        ("ALIGN",(1,0),(-1,-1),"CENTER"), ("ALIGN",(3,0),(3,-1),"RIGHT"),
        ("TOPPADDING",(0,0),(-1,-1),6), ("BOTTOMPADDING",(0,0),(-1,-1),6),
        ("LEFTPADDING",(0,0),(0,-1),10), ("GRID",(0,0),(-1,-1),0.4,BORDER),
    ]))
    story += [sl_tbl, Spacer(1, 0.15*inch)]

    # Provider summary (cumulative totals per provider)
    story.append(Paragraph("Provider Summary", section_style))
    prov_sum = df.groupby(["provider_name","provider_type","specialty","service_line"]).agg(
        Weeks=("week_ending","nunique"),
        Hours=("hours_worked","sum"),
        Days=("days_worked","sum"),
        Spend=("total_spend","sum")
    ).reset_index().sort_values("Spend", ascending=False)

    pv_hdr = ["Provider","Type","Specialty","Weeks","Hrs/Days","Total Spend"]
    pv_rows = [pv_hdr]
    for _, row in prov_sum.iterrows():
        hrs  = row["Hours"] if pd.notna(row["Hours"]) else 0
        days = row["Days"]  if pd.notna(row["Days"])  else 0
        if hrs > 0 and days > 0:   qty_str = f"{hrs:.1f}h/{days:.1f}d"
        elif hrs > 0:               qty_str = f"{hrs:.1f} hrs"
        else:                       qty_str = f"{days:.1f} days"
        pv_rows.append([row["provider_name"], row.get("provider_type",""),
                        row["specialty"], str(int(row["Weeks"])),
                        qty_str, f"${row['Spend']:,.2f}"])
    pv_tbl = Table(pv_rows, colWidths=[1.55*inch,0.65*inch,1.25*inch,0.55*inch,0.9*inch,1.0*inch])
    pv_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),BLUE), ("TEXTCOLOR",(0,0),(-1,0),WHITE),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"), ("FONTSIZE",(0,0),(-1,0),8),
        ("FONTNAME",(0,1),(-1,-1),"Helvetica"),     ("FONTSIZE",(0,1),(-1,-1),8),
        ("TEXTCOLOR",(0,1),(-1,-1),NAVY),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[WHITE,LIGHT]),
        ("ALIGN",(1,0),(-1,-1),"CENTER"), ("ALIGN",(5,0),(5,-1),"RIGHT"),
        ("TOPPADDING",(0,0),(-1,-1),6), ("BOTTOMPADDING",(0,0),(-1,-1),6),
        ("LEFTPADDING",(0,0),(0,-1),8), ("GRID",(0,0),(-1,-1),0.4,BORDER),
    ]))
    story += [pv_tbl]

    # Provider detail (individual entries) — optional
    if include_detail:
        story += [Spacer(1, 0.15*inch), Paragraph("Entry Detail", section_style)]
        det_hdr = [["Week","Provider","Type","Specialty","Hrs/Days","Rate","Total"]]
        if include_notes:
            det_hdr[0].append("Notes")
        det_rows = []
        for _, row in df.sort_values(["week_ending","specialty","provider_name"]).iterrows():
            sl = row.get("service_line","")
            if sl in HOURLY_SERVICE_LINES:
                qty  = f"{row['hours_worked']:.2f} hrs" if pd.notna(row.get("hours_worked")) else ""
                rate = f"${row['bill_rate']:,.2f}/hr"   if pd.notna(row.get("bill_rate"))    else ""
            else:
                qty  = f"{row['days_worked']:.1f} days" if pd.notna(row.get("days_worked")) else ""
                rate = f"${row['daily_rate']:,.2f}/day"  if pd.notna(row.get("daily_rate"))  else ""
            ot_flag = " (OT)" if row.get("is_ot") else ""
            r = [row["week_ending"],
                 row["provider_name"] + ot_flag,
                 row.get("provider_type",""),
                 row["specialty"], qty, rate, f"${row['total_spend']:,.2f}"]
            if include_notes:
                r.append(str(row.get("notes","") or ""))
            det_rows.append(r)
        det_cw = [0.8*inch,1.2*inch,0.55*inch,0.95*inch,0.65*inch,0.8*inch,0.7*inch,1.0*inch] if include_notes                  else [0.85*inch,1.4*inch,0.6*inch,1.1*inch,0.7*inch,0.85*inch,0.8*inch]
        det_tbl = Table(det_hdr+det_rows, colWidths=det_cw)
        det_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),BLUE), ("TEXTCOLOR",(0,0),(-1,0),WHITE),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"), ("FONTSIZE",(0,0),(-1,0),7),
            ("FONTNAME",(0,1),(-1,-1),"Helvetica"),     ("FONTSIZE",(0,1),(-1,-1),7),
            ("TEXTCOLOR",(0,1),(-1,-1),NAVY),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[WHITE,LIGHT]),
            ("ALIGN",(2,0),(-1,-1),"CENTER"), ("ALIGN",(6,0),(6,-1),"RIGHT"),
            ("TOPPADDING",(0,0),(-1,-1),4), ("BOTTOMPADDING",(0,0),(-1,-1),4),
            ("LEFTPADDING",(0,0),(0,-1),6), ("GRID",(0,0),(-1,-1),0.3,BORDER),
        ]))
        story.append(det_tbl)

    # Footer
    story += [Spacer(1,0.25*inch), HRFlowable(width="100%",thickness=0.5,color=BORDER), Spacer(1,0.08*inch)]
    ft = Table([[Paragraph(f"Confidential | {prepared_by}", small_style),
                 Paragraph(f"Report generated {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", right_style)]],
               colWidths=[3.5*inch, 3.5*inch])
    ft.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),
                             ("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),0)]))
    story.append(ft)

    doc.build(story)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
# PAGE CONFIG & CSS
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(page_title="RMCHCS Spend Tracker", page_icon="🏥",
                   layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=DM+Serif+Display&display=swap');
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    [data-testid="stSidebar"] { background: #0f1724; border-right: 1px solid #1e2d40; }
    [data-testid="stSidebar"] * { color: #cbd5e1 !important; }
    .main .block-container { background: #f8fafc; padding-top: 1.5rem; padding-bottom: 2rem; }
    .kpi-card { background:#fff; border-radius:12px; padding:1.4rem 1.6rem; border:1px solid #e2e8f0; box-shadow:0 1px 4px rgba(0,0,0,0.06); }
    .kpi-label { font-size:0.72rem; font-weight:600; letter-spacing:0.08em; text-transform:uppercase; color:#64748b; margin-bottom:0.3rem; }
    .kpi-value { font-family:'DM Serif Display',serif; font-size:2rem; color:#0f172a; line-height:1.1; }
    .kpi-delta-up { color:#10b981; font-weight:600; font-size:0.82rem; }
    .kpi-delta-neutral { color:#64748b; font-weight:500; font-size:0.82rem; }
    .section-header { font-family:'DM Serif Display',serif; font-size:1.25rem; color:#0f172a; margin-bottom:0.2rem; margin-top:0.5rem; }
    .section-sub { font-size:0.82rem; color:#64748b; margin-bottom:1rem; }
    .page-header { background:linear-gradient(135deg,#0f1724 0%,#1a3a5c 100%); border-radius:14px; padding:1.6rem 2rem; margin-bottom:1.5rem; display:flex; align-items:center; justify-content:space-between; }
    .page-header-title { font-family:'DM Serif Display',serif; font-size:1.6rem; color:#f1f5f9; margin:0; }
    .page-header-sub { font-size:0.82rem; color:#94a3b8; margin-top:0.2rem; }
    .page-header-badge { background:#1e3a5f; border:1px solid #2d5a8e; border-radius:8px; padding:0.5rem 1rem; font-size:0.78rem; color:#7dd3fc; font-weight:500; }
    .form-card { background:#fff; border-radius:12px; padding:1.5rem; border:1px solid #e2e8f0; box-shadow:0 1px 4px rgba(0,0,0,0.05); }
    [data-testid="metric-container"] { background:white; border-radius:12px; padding:1rem; border:1px solid #e2e8f0; }
    .stTabs [data-baseweb="tab-list"] { background:#f1f5f9; border-radius:10px; padding:4px; gap:2px; }
    .stTabs [data-baseweb="tab"] { border-radius:8px; font-weight:500; font-size:0.88rem; color:#64748b; }
    .stTabs [aria-selected="true"] { background:white !important; color:#0f172a !important; box-shadow:0 1px 3px rgba(0,0,0,0.1); }
    .stButton > button { border-radius:8px; font-weight:600; font-size:0.88rem; }
    #MainMenu {visibility:hidden;} footer {visibility:hidden;} header {visibility:hidden;}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# LOAD DATA & SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
df = load_data()

df_reqs  = load_reqs()
df_cands = load_candidates()

with st.sidebar:
    st.markdown("### 🏥 RMCHCS")
    st.markdown("**ITO Program Manager**")
    st.markdown("---")
    st.markdown("**💰 Spend Tracker**")
    page = st.radio("Navigate", [
        "Spend Dashboard", "Log Spend", "Manage Entries", "Generate Report",
        "─────────────",
        "Pipeline Dashboard", "Requisitions", "Candidates", "Credentialing",
        "─────────────",
        "Program ROI"
    ], index=0)
    st.markdown("---")
    if not df.empty:
        st.markdown("**Spend**")
        st.markdown(f"Total: **${df['total_spend'].sum():,.0f}**")
        st.markdown(f"Weeks: **{df['week_ending'].nunique()}**")
    if not df_reqs.empty:
        open_reqs = df_reqs[df_reqs["status"]=="Open"] if "status" in df_reqs.columns else pd.DataFrame()
        st.markdown("**Pipeline**")
        st.markdown(f"Open Reqs: **{len(open_reqs)}**")
        if not df_cands.empty:
            active = df_cands[df_cands["status"].isin(["Submitted","Clinical Call Scheduled",
                              "Clinical Call Complete","Offered","Accepted"])] if "status" in df_cands.columns else pd.DataFrame()
            st.markdown(f"Active Candidates: **{len(active)}**")
    st.markdown("---")
    st.markdown("<small style='color:#475569'>Ingenovis ITO<br>Internal Use Only</small>", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
if page == "Spend Dashboard":
    st.markdown("""
    <div class="page-header">
        <div>
            <div class="page-header-title">RMCHCS Spend Dashboard</div>
            <div class="page-header-sub">Rehoboth McKinley Christian Health Care Services · Gallup, NM</div>
        </div>
        <div class="page-header-badge">Ingenovis ITO</div>
    </div>""", unsafe_allow_html=True)

    if df.empty:
        st.info("No spend data yet. Go to **Log Spend** to add your first entry.")
        st.stop()

    col_f1, col_f2, col_f3 = st.columns(3)
    all_weeks = sorted(df["week_ending"].unique(), reverse=True)
    with col_f1:
        selected_week = st.selectbox("Filter by Week", ["All Weeks"] + list(all_weeks))
    with col_f2:
        selected_specialty = st.selectbox("Filter by Specialty", ["All Specialties"] + sorted(df["specialty"].unique().tolist()))
    with col_f3:
        selected_sl = st.selectbox("Filter by Service Line", ["All Service Lines"] + sorted(df["service_line"].unique().tolist()))

    filtered = df.copy()
    if selected_week != "All Weeks":      filtered = filtered[filtered["week_ending"] == selected_week]
    if selected_specialty != "All Specialties": filtered = filtered[filtered["specialty"] == selected_specialty]
    if selected_sl != "All Service Lines": filtered = filtered[filtered["service_line"] == selected_sl]

    total      = filtered["total_spend"].sum()
    providers  = filtered["provider_name"].nunique()
    specialties= filtered["specialty"].nunique()
    avg_rate   = filtered["daily_rate"].mean() if not filtered.empty else 0

    if len(all_weeks) >= 2 and selected_week == "All Weeks":
        tw = df[df["week_ending"]==all_weeks[0]]["total_spend"].sum()
        lw = df[df["week_ending"]==all_weeks[1]]["total_spend"].sum()
        dpct = ((tw-lw)/lw*100) if lw else 0
        dstr = f"{'▲' if dpct>=0 else '▼'} {abs(dpct):.1f}% vs last week"
        dcls = "kpi-delta-up" if dpct>=0 else "kpi-delta-neutral"
    else:
        dstr = f"{filtered['week_ending'].nunique()} week(s) shown"
        dcls = "kpi-delta-neutral"

    k1,k2,k3,k4 = st.columns(4)
    with k1:
        st.markdown(f'<div class="kpi-card"><div class="kpi-label">Total Spend</div><div class="kpi-value">${total:,.0f}</div><div class="{dcls}">{dstr}</div></div>', unsafe_allow_html=True)
    with k2:
        st.markdown(f'<div class="kpi-card"><div class="kpi-label">Active Providers</div><div class="kpi-value">{providers}</div><div class="kpi-delta-neutral">Unique in view</div></div>', unsafe_allow_html=True)
    with k3:
        st.markdown(f'<div class="kpi-card"><div class="kpi-label">Specialties Covered</div><div class="kpi-value">{specialties}</div><div class="kpi-delta-neutral">Across all service lines</div></div>', unsafe_allow_html=True)
    with k4:
        st.markdown(f'<div class="kpi-card"><div class="kpi-label">Avg Daily Rate</div><div class="kpi-value">${avg_rate:,.0f}</div><div class="kpi-delta-neutral">Per provider per day</div></div>', unsafe_allow_html=True)

    st.markdown("<div style='margin-top:1.5rem'></div>", unsafe_allow_html=True)

    # Row 1: weekly trend + specialty donut
    c1, c2 = st.columns([3,2])
    with c1:
        st.markdown('<div class="section-header">Weekly Spend Trend</div>', unsafe_allow_html=True)
        st.markdown('<div class="section-sub">Total spend logged per week ending date</div>', unsafe_allow_html=True)
        weekly = df.groupby("week_ending")["total_spend"].sum().reset_index().sort_values("week_ending")
        # Use string labels so Plotly treats weeks as discrete categories, not a continuous date axis
        weekly["week_label"] = weekly["week_ending"].astype(str)
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=weekly["week_label"], y=weekly["total_spend"],
            marker_color="#1a3a5c", marker_line_width=0,
            hovertemplate="<b>Week Ending:</b> %{x}<br><b>Spend:</b> $%{y:,.0f}<extra></extra>"
        ))
        fig.add_trace(go.Scatter(
            x=weekly["week_label"], y=weekly["total_spend"], mode="lines+markers",
            line=dict(color="#3b82f6", width=2.5), marker=dict(size=7, color="#3b82f6"),
            hoverinfo="skip"
        ))
        y_max = weekly["total_spend"].max() * 2.0 if not weekly.empty else 1000
        fig.update_layout(
            plot_bgcolor="white", paper_bgcolor="white",
            margin=dict(l=0,r=0,t=10,b=0), height=280,
            bargap=0.35,
            xaxis=dict(showgrid=False, tickfont=dict(size=11), type="category"),
            yaxis=dict(showgrid=True, gridcolor="#f1f5f9", tickprefix="$",
                       tickfont=dict(size=11), range=[0, y_max]),
            showlegend=False
        )
        st.plotly_chart(fig, use_container_width=True)

    with c2:
        st.markdown('<div class="section-header">Spend by Specialty</div>', unsafe_allow_html=True)
        st.markdown('<div class="section-sub">Share of total spend per specialty</div>', unsafe_allow_html=True)
        spec_spend = filtered.groupby("specialty")["total_spend"].sum().reset_index().sort_values("total_spend", ascending=False)
        fig2 = go.Figure(go.Pie(labels=spec_spend["specialty"], values=spec_spend["total_spend"], hole=0.52,
                                 marker=dict(colors=["#0f1724","#1a3a5c","#2d5a8e","#3b82f6","#60a5fa","#93c5fd","#bfdbfe"][:len(spec_spend)]),
                                 textinfo="percent",
                                 hovertemplate="<b>%{label}</b><br>$%{value:,.0f}<br>%{percent}<extra></extra>"))
        fig2.update_layout(plot_bgcolor="white", paper_bgcolor="white", margin=dict(l=0,r=0,t=10,b=0), height=280,
                           legend=dict(font=dict(size=11)))
        st.plotly_chart(fig2, use_container_width=True)

    # Row 2: provider bar + service line bar
    c3, c4 = st.columns([3,2])
    with c3:
        st.markdown('<div class="section-header">Spend by Provider</div>', unsafe_allow_html=True)
        st.markdown('<div class="section-sub">Total spend per provider in current view</div>', unsafe_allow_html=True)
        prov_spend = filtered.groupby("provider_name")["total_spend"].sum().reset_index().sort_values("total_spend")
        fig3 = go.Figure(go.Bar(x=prov_spend["total_spend"], y=prov_spend["provider_name"], orientation="h",
                                 marker_color="#2d5a8e", marker_line_width=0,
                                 hovertemplate="<b>%{y}</b><br>$%{x:,.0f}<extra></extra>"))
        fig3.update_layout(plot_bgcolor="white", paper_bgcolor="white", margin=dict(l=0,r=10,t=10,b=0),
                           height=max(200, len(prov_spend)*38),
                           xaxis=dict(showgrid=True, gridcolor="#f1f5f9", tickprefix="$", tickfont=dict(size=11)),
                           yaxis=dict(showgrid=False, tickfont=dict(size=11)))
        st.plotly_chart(fig3, use_container_width=True)

    with c4:
        st.markdown('<div class="section-header">Service Line Breakdown</div>', unsafe_allow_html=True)
        st.markdown('<div class="section-sub">Physician vs APP vs other lines</div>', unsafe_allow_html=True)
        sl_spend = filtered.groupby("service_line")["total_spend"].sum().reset_index().sort_values("total_spend", ascending=False)
        fig4 = go.Figure(go.Bar(x=sl_spend["service_line"], y=sl_spend["total_spend"],
                                 marker_color=["#1a3a5c","#3b82f6","#60a5fa","#93c5fd"][:len(sl_spend)],
                                 marker_line_width=0,
                                 hovertemplate="<b>%{x}</b><br>$%{y:,.0f}<extra></extra>"))
        fig4.update_layout(plot_bgcolor="white", paper_bgcolor="white", margin=dict(l=0,r=0,t=10,b=0), height=280,
                           xaxis=dict(showgrid=False, tickfont=dict(size=11)),
                           yaxis=dict(showgrid=True, gridcolor="#f1f5f9", tickprefix="$", tickfont=dict(size=11)))
        st.plotly_chart(fig4, use_container_width=True)

    st.markdown('<div class="section-header" style="margin-top:0.5rem">Entry Detail</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-sub">All logged entries in current filter view</div>', unsafe_allow_html=True)
    disp = filtered[["week_ending","provider_name","specialty","service_line",
                       "days_worked","daily_rate","hours_worked","bill_rate",
                       "total_spend","notes"]].sort_values("week_ending", ascending=False).copy()
    def fmt_unit(row):
        if row["service_line"] in HOURLY_SERVICE_LINES:
            hrs  = row["hours_worked"] if pd.notna(row["hours_worked"]) else ""
            rate = f"${row['bill_rate']:,.2f}/hr" if pd.notna(row["bill_rate"]) else ""
            return pd.Series([hrs, rate])
        else:
            days = row["days_worked"] if pd.notna(row["days_worked"]) else ""
            rate = f"${row['daily_rate']:,.2f}/day" if pd.notna(row["daily_rate"]) else ""
            return pd.Series([days, rate])
    disp[["_qty","_rate"]] = disp.apply(fmt_unit, axis=1)
    disp["total_spend"] = disp["total_spend"].apply(lambda x: f"${x:,.2f}" if pd.notna(x) else "")
    if "is_ot" not in disp.columns: disp["is_ot"] = False
    disp["_ot"] = disp["is_ot"].apply(lambda x: "OT" if x else "Regular")
    disp_final = disp[["week_ending","provider_name","specialty","service_line","_ot","_qty","_rate","total_spend","notes"]].copy()
    disp_final.columns = ["Week Ending","Provider","Specialty","Service Line","Type","Days / Hours","Rate","Total Spend","Notes"]
    st.dataframe(disp_final, use_container_width=True, hide_index=True)

# ══════════════════════════════════════════════════════════════════════════════
# LOG SPEND
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Log Spend":
    st.markdown("""
    <div class="page-header">
        <div>
            <div class="page-header-title">Log Weekly Spend</div>
            <div class="page-header-sub">Add a single entry manually or upload multiple rows at once</div>
        </div>
    </div>""", unsafe_allow_html=True)

    tab_manual, tab_bulk, tab_invoice = st.tabs(["Single Entry", "Bulk Upload (CSV)", "Upload Invoice (Excel)"])

    with tab_manual:
        st.markdown("<div style='height:0.75rem'></div>", unsafe_allow_html=True)
        specialty_options = ["Emergency Medicine","Hospitalist/Internal Medicine","Family Medicine",
                             "OB/GYN","Surgery","Radiology","Anesthesiology","Psychiatry",
                             "Pediatrics","Cardiology","Orthopedics","Neurology","Other"]
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Week & Provider**")
            week_ending   = st.date_input("Week Ending Date", value=get_week_ending())
            provider_name = st.text_input("Provider Name", placeholder="Dr. Jane Smith")
            provider_type = st.selectbox("Provider Type", ["Physician","APP","CRNA","NP","PA","RN","LPN","Tech","Other"])
        with col2:
            st.markdown("**Assignment Details**")
            specialty    = st.selectbox("Specialty", specialty_options)
            service_line = st.selectbox("Service Line", ["Physician","APP","Nursing","Allied Health","Other"])
            department   = st.text_input("Department / Unit", placeholder="e.g., ED, ICU, L&D")

        st.markdown("---")
        is_hourly = service_line in HOURLY_SERVICE_LINES

        if is_hourly:
            st.markdown("**Billing: Hourly** (Nursing / Allied Health)")
            c3,c4,c5 = st.columns(3)
            with c3: hours_worked = st.number_input("Hours Worked", min_value=0.5, value=40.0, step=0.5)
            with c4: bill_rate    = st.number_input("Bill Rate ($/hr)", min_value=0.0, value=75.0, step=1.0)
            with c5: st.metric("Calculated Total", f"${hours_worked*bill_rate:,.2f}")
            days_worked = None; daily_rate = None
            calc_total  = round(hours_worked * bill_rate, 2)
        else:
            st.markdown("**Billing: Daily** (Physician / APP / NP / PA)")
            c3,c4,c5 = st.columns(3)
            with c3: days_worked = st.number_input("Days Worked", min_value=0.5, max_value=7.0, value=5.0, step=0.5)
            with c4: daily_rate  = st.number_input("Daily Rate ($)", min_value=0.0, value=1800.0, step=50.0)
            with c5: st.metric("Calculated Total", f"${days_worked*daily_rate:,.2f}")
            hours_worked = None; bill_rate = None
            calc_total   = round(days_worked * daily_rate, 2)

        notes = st.text_area("Notes (optional)", placeholder="Contract type, extension, specialty notes...", height=80)
        st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)
        cb1, _ = st.columns([1,4])
        with cb1:
            if st.button("Save Entry", type="primary", use_container_width=True):
                if not provider_name.strip():
                    st.error("Provider name is required.")
                else:
                    save_entry({
                        "week_ending":  str(week_ending),
                        "provider_name":provider_name.strip(),
                        "provider_type":provider_type,
                        "specialty":    specialty,
                        "service_line": service_line,
                        "department":   department.strip(),
                        "days_worked":  days_worked,
                        "daily_rate":   daily_rate,
                        "hours_worked": hours_worked,
                        "bill_rate":    bill_rate,
                        "total_spend":  calc_total,
                        "notes":        notes.strip(),
                        "logged_at":    datetime.now().isoformat()
                    })
                    st.success(f"Saved. {provider_name} | {week_ending} | ${calc_total:,.2f}")
                    st.rerun()

    with tab_bulk:
        st.markdown("<div style='height:0.75rem'></div>", unsafe_allow_html=True)
        st.markdown('<div class="section-header">Step 1: Download the Template</div>', unsafe_allow_html=True)
        st.markdown('<div class="section-sub">Fill this out in Excel or Google Sheets. Do not rename the column headers.</div>', unsafe_allow_html=True)
        template_df = pd.DataFrame({
            "week_ending":  ["2025-06-07","2025-06-07","2025-06-07"],
            "provider_name":["Dr. Jane Smith","Sarah Jones NP","Mary RN"],
            "provider_type":["Physician","NP","RN"],
            "specialty":    ["Emergency Medicine","Family Medicine","Med/Surg"],
            "service_line": ["Physician","APP","Nursing"],
            "department":   ["ED","Clinic","3 West"],
            "days_worked":  [5,3,""],
            "daily_rate":   [1800,950,""],
            "hours_worked": ["","",36],
            "bill_rate":    ["","",75],
            "notes":        ["Extended contract","",""]
        })
        st.download_button("Download CSV Template", template_df.to_csv(index=False).encode("utf-8"),
                           "RMCHCS_SpendUpload_Template.csv", "text/csv", type="primary")
        st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)
        st.markdown('<div class="section-header">Step 2: Upload Your Filled File</div>', unsafe_allow_html=True)
        st.markdown('<div class="section-sub">CSV only. Total is calculated automatically from days x rate.</div>', unsafe_allow_html=True)
        uploaded = st.file_uploader("Choose CSV", type=["csv"], label_visibility="collapsed")
        if uploaded:
            try:
                udf = pd.read_csv(uploaded)
                required = ["week_ending","provider_name","provider_type","specialty","service_line"]
                missing  = [c for c in required if c not in udf.columns]
                if missing:
                    st.error(f"Missing columns: {', '.join(missing)}. Use the template above.")
                else:
                    for col in ["department","notes","days_worked","daily_rate","hours_worked","bill_rate"]:
                        if col not in udf.columns: udf[col] = ""
                    for col in ["days_worked","daily_rate","hours_worked","bill_rate"]:
                        udf[col] = pd.to_numeric(udf[col], errors="coerce")

                    def calc_row_total(row):
                        sl = str(row.get("service_line",""))
                        if sl in HOURLY_SERVICE_LINES:
                            if pd.notna(row["hours_worked"]) and pd.notna(row["bill_rate"]):
                                return round(row["hours_worked"] * row["bill_rate"], 2)
                        else:
                            if pd.notna(row["days_worked"]) and pd.notna(row["daily_rate"]):
                                return round(row["days_worked"] * row["daily_rate"], 2)
                        return None

                    udf["total_spend"] = udf.apply(calc_row_total, axis=1)
                    bad   = udf[udf["total_spend"].isnull()]
                    clean = udf[udf["total_spend"].notna()].copy()

                    st.markdown('<div class="section-header">Step 3: Preview Before Saving</div>', unsafe_allow_html=True)
                    if not bad.empty:
                        st.warning(f"{len(bad)} row(s) skipped. Physician/APP rows need days_worked + daily_rate. Nursing/Allied rows need hours_worked + bill_rate.")
                    if clean.empty:
                        st.error("No valid rows to import.")
                    else:
                        def fmt_prev_row(row):
                            sl = str(row.get("service_line",""))
                            if sl in HOURLY_SERVICE_LINES:
                                qty  = f"{row['hours_worked']:.1f} hrs" if pd.notna(row["hours_worked"]) else ""
                                rate = f"${row['bill_rate']:,.2f}/hr"   if pd.notna(row["bill_rate"])    else ""
                            else:
                                qty  = f"{row['days_worked']:.1f} days" if pd.notna(row["days_worked"]) else ""
                                rate = f"${row['daily_rate']:,.2f}/day"  if pd.notna(row["daily_rate"])  else ""
                            return pd.Series([qty, rate])
                        prev = clean.copy()
                        prev[["_qty","_rate"]] = prev.apply(fmt_prev_row, axis=1)
                        prev["total_spend"] = prev["total_spend"].apply(lambda x: f"${x:,.2f}")
                        prev_show = prev[["week_ending","provider_name","provider_type","specialty","service_line","_qty","_rate","total_spend","notes"]].copy()
                        prev_show.columns = ["Week Ending","Provider","Type","Specialty","Service Line","Qty","Rate","Total Spend","Notes"]
                        st.dataframe(prev_show, use_container_width=True, hide_index=True)
                        bulk_total = clean["total_spend"].sum()
                        st.markdown(f"**{len(clean)} rows ready | Combined total: ${bulk_total:,.2f}**")
                        cb2, _ = st.columns([1,4])
                        with cb2:
                            if st.button("Save All Entries", type="primary", use_container_width=True):
                                for _, row in clean.iterrows():
                                    sl = str(row.get("service_line",""))
                                    save_entry({
                                        "week_ending":  str(row["week_ending"]).strip(),
                                        "provider_name":str(row["provider_name"]).strip(),
                                        "provider_type":str(row.get("provider_type","")).strip(),
                                        "specialty":    str(row["specialty"]).strip(),
                                        "service_line": sl,
                                        "department":   str(row.get("department","")).strip(),
                                        "days_worked":  float(row["days_worked"]) if pd.notna(row.get("days_worked")) else None,
                                        "daily_rate":   float(row["daily_rate"])  if pd.notna(row.get("daily_rate"))  else None,
                                        "hours_worked": float(row["hours_worked"])if pd.notna(row.get("hours_worked"))else None,
                                        "bill_rate":    float(row["bill_rate"])   if pd.notna(row.get("bill_rate"))   else None,
                                        "total_spend":  float(row["total_spend"]),
                                        "notes":        str(row.get("notes","")).strip(),
                                        "logged_at":    datetime.now().isoformat()
                                    })
                                st.success(f"{len(clean)} entries saved. Total: ${bulk_total:,.2f}")
                                st.rerun()
            except Exception as e:
                st.error(f"Could not read file: {e}")

    with tab_invoice:
        st.markdown("<div style='height:0.75rem'></div>", unsafe_allow_html=True)
        st.markdown('<div class="section-header">Upload Vista Consolidated Invoice</div>', unsafe_allow_html=True)
        st.markdown('<div class="section-sub">Drop your Excel invoice file here. Regular and OT lines are imported as separate entries. Specialty, service line, and provider type are auto-detected from the Placement field.</div>', unsafe_allow_html=True)

        inv_file = st.file_uploader("Choose Excel invoice (.xlsx)", type=["xlsx"], label_visibility="collapsed")

        if inv_file:
            with st.spinner("Parsing invoice..."):
                try:
                    parsed_rows, parse_warnings, skipped_count = parse_invoice_xlsx(inv_file)
                except Exception as e:
                    st.error(f"Could not read invoice file: {e}")
                    parsed_rows = []
                    parse_warnings = []
                    skipped_count = 0

            if parse_warnings:
                for w in parse_warnings:
                    st.warning(w)

            if not parsed_rows:
                st.error("No valid rows found in this file.")
            else:
                inv_df = pd.DataFrame(parsed_rows)

                st.markdown(f"<div style='height:0.5rem'></div>", unsafe_allow_html=True)
                st.markdown('<div class="section-header">Preview</div>', unsafe_allow_html=True)

                # Summary KPIs
                ki1, ki2, ki3, ki4 = st.columns(4)
                ki1.metric("Rows Found",       len(inv_df))
                ki2.metric("Providers",         inv_df["provider_name"].nunique())
                ki3.metric("Weeks",             inv_df["week_ending"].nunique())
                ki4.metric("Total Spend",       f"${inv_df['total_spend'].sum():,.2f}")

                if skipped_count:
                    st.caption(f"{skipped_count} summary/blank rows skipped automatically.")

                # Preview table
                prev_inv = inv_df.copy()
                prev_inv["is_ot"]       = prev_inv["is_ot"].apply(lambda x: "OT" if x else "Regular")
                prev_inv["hours_worked"]= prev_inv["hours_worked"].apply(lambda x: f"{x:.2f} hrs")
                prev_inv["bill_rate"]   = prev_inv["bill_rate"].apply(lambda x: f"${x:,.2f}/hr")
                prev_inv["total_spend"] = prev_inv["total_spend"].apply(lambda x: f"${x:,.2f}")
                st.dataframe(
                    prev_inv[["week_ending","provider_name","provider_type","specialty",
                               "service_line","is_ot","hours_worked","bill_rate","total_spend","invoice_number"]].rename(
                        columns={"week_ending":"Week Ending","provider_name":"Provider",
                                 "provider_type":"Type","specialty":"Specialty",
                                 "service_line":"Service Line","is_ot":"Rate Type",
                                 "hours_worked":"Hours","bill_rate":"Bill Rate",
                                 "total_spend":"Total Spend","invoice_number":"Invoice #"}),
                    use_container_width=True, hide_index=True
                )

                st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)
                st.markdown("**Review auto-detected fields below. Edit anything before saving.**")

                # Editable corrections
                with st.expander("Edit specialty / service line / provider type before saving", expanded=False):
                    st.caption("These are auto-detected from the invoice Placement field. Correct any that are wrong.")
                    providers_to_edit = inv_df["provider_name"].unique().tolist()
                    corrections = {}
                    for pname in providers_to_edit:
                        prows = inv_df[inv_df["provider_name"] == pname].iloc[0]
                        st.markdown(f"**{pname}**")
                        ec1, ec2, ec3 = st.columns(3)
                        corrections[pname] = {
                            "provider_type": ec1.text_input(
                                "Provider Type", value=prows["provider_type"],
                                key=f"pt_{pname}"),
                            "specialty": ec2.text_input(
                                "Specialty", value=prows["specialty"],
                                key=f"sp_{pname}"),
                            "service_line": ec3.selectbox(
                                "Service Line",
                                ["Allied Health","Nursing","Physician","APP","Other"],
                                index=["Allied Health","Nursing","Physician","APP","Other"].index(
                                    prows["service_line"]) if prows["service_line"] in
                                    ["Allied Health","Nursing","Physician","APP","Other"] else 0,
                                key=f"sl_{pname}"),
                            "department": ec1.text_input(
                                "Department", value="",
                                key=f"dept_{pname}")
                        }

                st.markdown("<div style='height:0.4rem'></div>", unsafe_allow_html=True)
                ci1, _ = st.columns([1,4])
                with ci1:
                    if st.button("Save All Invoice Entries", type="primary", use_container_width=True):
                        saved = 0
                        for _, row in inv_df.iterrows():
                            pname = row["provider_name"]
                            cor   = corrections.get(pname, {})
                            save_entry({
                                "week_ending":    row["week_ending"],
                                "provider_name":  pname,
                                "provider_type":  cor.get("provider_type", row["provider_type"]),
                                "specialty":      cor.get("specialty",     row["specialty"]),
                                "service_line":   cor.get("service_line",  row["service_line"]),
                                "department":     cor.get("department",    ""),
                                "days_worked":    None,
                                "daily_rate":     None,
                                "hours_worked":   row["hours_worked"],
                                "bill_rate":      row["bill_rate"],
                                "is_ot":          bool(row["is_ot"]),
                                "invoice_number": row["invoice_number"],
                                "total_spend":    row["total_spend"],
                                "notes":          row["notes"],
                                "logged_at":      datetime.now().isoformat()
                            })
                            saved += 1
                        st.success(f"{saved} entries saved from invoice. Total: ${inv_df['total_spend'].sum():,.2f}")
                        st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# MANAGE ENTRIES
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Manage Entries":
    st.markdown("""
    <div class="page-header">
        <div>
            <div class="page-header-title">Manage Entries</div>
            <div class="page-header-sub">Review and delete logged spend entries</div>
        </div>
    </div>""", unsafe_allow_html=True)

    if df.empty:
        st.info("No entries logged yet.")
        st.stop()

    all_weeks = sorted(df["week_ending"].unique(), reverse=True)
    sel_week  = st.selectbox("Select Week", ["All Weeks"] + list(all_weeks))
    view_df   = df.copy() if sel_week=="All Weeks" else df[df["week_ending"]==sel_week].copy()
    view_df   = view_df.sort_values("week_ending", ascending=False).reset_index(drop=True)
    st.markdown(f"**{len(view_df)} entries** in current view")

    for _, row in view_df.iterrows():
        with st.expander(f"{row['week_ending']} | {row['provider_name']} | {row['specialty']} | ${row['total_spend']:,.2f}"):
            c1,c2,c3 = st.columns(3)
            c1.write(f"**Provider Type:** {row.get('provider_type','N/A')}")
            c2.write(f"**Service Line:** {row['service_line']}")
            c3.write(f"**Department:** {row.get('department','N/A')}")
            sl = row.get("service_line","")
            if sl in HOURLY_SERVICE_LINES:
                hrs  = row["hours_worked"] if pd.notna(row.get("hours_worked")) else "N/A"
                rate = f"${row['bill_rate']:,.2f}/hr" if pd.notna(row.get("bill_rate")) else "N/A"
                c1.write(f"**Hours Worked:** {hrs}")
                c2.write(f"**Bill Rate:** {rate}")
            else:
                days = row["days_worked"] if pd.notna(row.get("days_worked")) else "N/A"
                rate = f"${row['daily_rate']:,.2f}/day" if pd.notna(row.get("daily_rate")) else "N/A"
                c1.write(f"**Days Worked:** {days}")
                c2.write(f"**Daily Rate:** {rate}")
            c3.write(f"**Total Spend:** ${row['total_spend']:,.2f}")
            if row.get("notes"): st.write(f"**Notes:** {row['notes']}")
            if row.get("invoice_number"): st.write(f"**Invoice #:** {row['invoice_number']}")
            ot_label = " 🔶 OT" if row.get("is_ot") else ""
            if ot_label: st.caption(f"Overtime entry{ot_label}")
            if st.button("Delete Entry", key=f"del_{row['id']}", type="secondary"):
                delete_entry(row["id"])
                st.success("Entry deleted.")
                st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# GENERATE REPORT
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Generate Report":
    st.markdown("""
    <div class="page-header">
        <div>
            <div class="page-header-title">Generate Client Report</div>
            <div class="page-header-sub">Weekly snapshot or cumulative total spend report</div>
        </div>
    </div>""", unsafe_allow_html=True)

    if df.empty:
        st.info("No data available to report on yet.")
        st.stop()

    tab_weekly, tab_cumulative = st.tabs(["Weekly Report", "Total Spend Report"])

    # ── Weekly ──────────────────────────────────────────────────────────────
    with tab_weekly:
        st.markdown("<div style='height:0.75rem'></div>", unsafe_allow_html=True)
        all_weeks = sorted(df["week_ending"].unique(), reverse=True)
        c1,c2 = st.columns(2)
        with c1:
            report_week  = st.selectbox("Select Week to Report", all_weeks)
            report_title = st.text_input("Report Title", value=f"Weekly Spend Report - Week Ending {report_week}")
        with c2:
            include_detail = st.checkbox("Include Provider Detail Table", value=True)
            include_notes  = st.checkbox("Include Notes Column", value=True)
            prepared_by    = st.text_input("Prepared By", value="Ingenovis ITO")

        st.markdown("---")
        week_data = df[df["week_ending"]==report_week]
        m1,m2,m3 = st.columns(3)
        m1.metric("Week Total Spend", f"${week_data['total_spend'].sum():,.2f}")
        m2.metric("Providers",  week_data["provider_name"].nunique())
        m3.metric("Specialties",week_data["specialty"].nunique())

        def fmt_report_row(row):
            sl = row.get("service_line","")
            if sl in HOURLY_SERVICE_LINES:
                qty  = f"{row['hours_worked']:.1f} hrs" if pd.notna(row.get("hours_worked")) else ""
                rate = f"${row['bill_rate']:,.2f}/hr"   if pd.notna(row.get("bill_rate"))    else ""
            else:
                qty  = f"{row['days_worked']:.1f} days" if pd.notna(row.get("days_worked")) else ""
                rate = f"${row['daily_rate']:,.2f}/day"  if pd.notna(row.get("daily_rate"))  else ""
            return pd.Series([qty, rate])

        rpt_disp = week_data.copy()
        rpt_disp[["_qty","_rate"]] = rpt_disp.apply(fmt_report_row, axis=1)
        rpt_disp["total_spend"] = rpt_disp["total_spend"].apply(lambda x: f"${x:,.2f}")
        st.dataframe(rpt_disp[["provider_name","specialty","service_line","_qty","_rate","total_spend"]].rename(
            columns={"provider_name":"Provider","specialty":"Specialty","service_line":"Service Line",
                     "_qty":"Days / Hours","_rate":"Rate","total_spend":"Total Spend"}),
            use_container_width=True, hide_index=True)

        st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)
        cg1, _ = st.columns([1,4])
        with cg1:
            if st.button("Generate Weekly PDF", type="primary", use_container_width=True):
                with st.spinner("Building report..."):
                    pdf_bytes = generate_pdf_report(week_data, report_week, report_title,
                                                     prepared_by, include_detail, include_notes)
                st.download_button("Download Weekly PDF", pdf_bytes,
                                   f"RMCHCS_SpendReport_{report_week}.pdf", "application/pdf", type="primary")
                st.success("Report ready. Click above to download.")

    # ── Cumulative ──────────────────────────────────────────────────────────
    with tab_cumulative:
        st.markdown("<div style='height:0.75rem'></div>", unsafe_allow_html=True)
        all_weeks_sorted = sorted(df["week_ending"].unique())
        c1,c2 = st.columns(2)
        with c1:
            date_from = st.selectbox("From Week Ending", all_weeks_sorted,
                                     index=0)
            date_to   = st.selectbox("To Week Ending",
                                     [w for w in all_weeks_sorted if w >= date_from],
                                     index=len([w for w in all_weeks_sorted if w >= date_from])-1)
            cum_title = st.text_input("Report Title",
                                      value=f"RMCHCS Total Spend Report — {date_from} to {date_to}",
                                      key="cum_title")
        with c2:
            cum_detail  = st.checkbox("Include Entry Detail Table", value=False,
                                       help="Shows every individual entry — can get long for multi-week reports")
            cum_notes   = st.checkbox("Include Notes Column", value=True, key="cum_notes")
            cum_prep_by = st.text_input("Prepared By", value="Ingenovis ITO", key="cum_prep")

        st.markdown("---")
        cum_data = df[(df["week_ending"] >= date_from) & (df["week_ending"] <= date_to)]

        if cum_data.empty:
            st.warning("No data in selected date range.")
        else:
            ka1,ka2,ka3,ka4 = st.columns(4)
            ka1.metric("Total Spend",    f"${cum_data['total_spend'].sum():,.2f}")
            ka2.metric("Providers",       cum_data["provider_name"].nunique())
            ka3.metric("Weeks",           cum_data["week_ending"].nunique())
            ka4.metric("Specialties",     cum_data["specialty"].nunique())

            st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)
            st.markdown("**Weekly Breakdown**")
            weekly_prev = cum_data.groupby("week_ending").agg(
                Providers=("provider_name","nunique"),
                Spend=("total_spend","sum")
            ).reset_index().sort_values("week_ending")
            weekly_prev["Total Spend"] = weekly_prev["Spend"].apply(lambda x: f"${x:,.2f}")
            running = 0
            cumulative_vals = []
            for s in weekly_prev["Spend"]:
                running += s
                cumulative_vals.append(f"${running:,.2f}")
            weekly_prev["Cumulative"] = cumulative_vals
            st.dataframe(
                weekly_prev[["week_ending","Providers","Total Spend","Cumulative"]].rename(
                    columns={"week_ending":"Week Ending","Providers":"Providers"}),
                use_container_width=True, hide_index=True)

            st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)
            cg2, _ = st.columns([1,4])
            with cg2:
                if st.button("Generate Total Spend PDF", type="primary", use_container_width=True):
                    with st.spinner("Building report..."):
                        pdf_bytes = generate_cumulative_report(
                            cum_data, date_from, date_to, cum_title,
                            cum_prep_by, cum_detail, cum_notes)
                    st.download_button("Download Total Spend PDF", pdf_bytes,
                                       f"RMCHCS_TotalSpend_{date_from}_to_{date_to}.pdf",
                                       "application/pdf", type="primary")
                    st.success("Report ready. Click above to download.")


# ══════════════════════════════════════════════════════════════════════════════
# PIPELINE DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Pipeline Dashboard":
    st.markdown("""
    <div class="page-header">
        <div>
            <div class="page-header-title">Pipeline Dashboard</div>
            <div class="page-header-sub">Requisitions, candidates, and credentialing at a glance</div>
        </div>
        <div class="page-header-badge">Ingenovis ITO</div>
    </div>""", unsafe_allow_html=True)

    if df_reqs.empty and df_cands.empty:
        st.info("No pipeline data yet. Start by adding requisitions and candidates.")
        st.stop()

    # KPI row
    total_reqs   = len(df_reqs)
    open_reqs    = len(df_reqs[df_reqs["status"]=="Open"]) if not df_reqs.empty else 0
    total_cands  = len(df_cands)
    placed       = len(df_cands[df_cands["status"]=="Placed"]) if not df_cands.empty else 0
    active_pipe  = len(df_cands[df_cands["status"].isin(["Submitted","Clinical Call Scheduled",
                       "Clinical Call Complete","Offered","Accepted"])]) if not df_cands.empty else 0
    cred_due     = 0
    if not df_cands.empty and "cred_due_date" in df_cands.columns:
        today_str = date.today().isoformat()
        cred_due = len(df_cands[
            (df_cands["cred_due_date"].notna()) &
            (df_cands["cred_due_date"] != "") &
            (df_cands["cred_due_date"] <= today_str) &
            (~df_cands["cred_status"].isin(["Clear","Cancelled"]))
        ]) if "cred_status" in df_cands.columns else 0

    k1,k2,k3,k4 = st.columns(4)
    with k1:
        st.markdown(f'<div class="kpi-card"><div class="kpi-label">Open Requisitions</div><div class="kpi-value">{open_reqs}</div><div class="kpi-delta-neutral">{total_reqs} total reqs</div></div>', unsafe_allow_html=True)
    with k2:
        st.markdown(f'<div class="kpi-card"><div class="kpi-label">Active Pipeline</div><div class="kpi-value">{active_pipe}</div><div class="kpi-delta-neutral">candidates in process</div></div>', unsafe_allow_html=True)
    with k3:
        st.markdown(f'<div class="kpi-card"><div class="kpi-label">Placed</div><div class="kpi-value">{placed}</div><div class="kpi-delta-neutral">of {total_cands} submitted</div></div>', unsafe_allow_html=True)
    with k4:
        dcls = "kpi-delta-up" if cred_due == 0 else "kpi-delta-neutral"
        st.markdown(f'<div class="kpi-card"><div class="kpi-label">Cred Due / Overdue</div><div class="kpi-value">{cred_due}</div><div class="{dcls}">{"All clear" if cred_due==0 else "needs attention"}</div></div>', unsafe_allow_html=True)

    st.markdown("<div style='margin-top:1.5rem'></div>", unsafe_allow_html=True)

    col_l, col_r = st.columns(2)

    with col_l:
        st.markdown('<div class="section-header">Requisitions by Status</div>', unsafe_allow_html=True)
        if not df_reqs.empty:
            req_status = df_reqs.groupby("status").size().reset_index(name="count")
            fig_rs = go.Figure(go.Bar(
                x=req_status["status"], y=req_status["count"],
                marker_color=["#1a3a5c" if s=="Open" else "#3b82f6" if s=="On Hold"
                              else "#60a5fa" if s=="Max Submissions" else "#10b981" if s=="Filled"
                              else "#94a3b8" for s in req_status["status"]],
                marker_line_width=0,
                hovertemplate="<b>%{x}</b><br>%{y} reqs<extra></extra>"
            ))
            fig_rs.update_layout(plot_bgcolor="white", paper_bgcolor="white",
                                  margin=dict(l=0,r=0,t=10,b=0), height=220,
                                  xaxis=dict(showgrid=False, tickfont=dict(size=11)),
                                  yaxis=dict(showgrid=True, gridcolor="#f1f5f9", tickfont=dict(size=11), dtick=1))
            st.plotly_chart(fig_rs, use_container_width=True)
        else:
            st.info("No requisitions logged yet.")

    with col_r:
        st.markdown('<div class="section-header">Candidates by Status</div>', unsafe_allow_html=True)
        if not df_cands.empty:
            cand_status = df_cands.groupby("status").size().reset_index(name="count")
            fig_cs = go.Figure(go.Bar(
                x=cand_status["status"], y=cand_status["count"],
                marker_color="#2d5a8e", marker_line_width=0,
                hovertemplate="<b>%{x}</b><br>%{y} candidates<extra></extra>"
            ))
            fig_cs.update_layout(plot_bgcolor="white", paper_bgcolor="white",
                                  margin=dict(l=0,r=0,t=10,b=0), height=220,
                                  xaxis=dict(showgrid=False, tickfont=dict(size=10), tickangle=-20),
                                  yaxis=dict(showgrid=True, gridcolor="#f1f5f9", tickfont=dict(size=11), dtick=1))
            st.plotly_chart(fig_cs, use_container_width=True)
        else:
            st.info("No candidates logged yet.")

    # Open reqs table
    if not df_reqs.empty:
        st.markdown('<div class="section-header" style="margin-top:0.5rem">Open Requisitions</div>', unsafe_allow_html=True)
        open_df = df_reqs[df_reqs["status"]=="Open"].copy() if "status" in df_reqs.columns else df_reqs.copy()
        if open_df.empty:
            st.info("No open requisitions.")
        else:
            disp_cols = [c for c in ["specialty","job_title","discipline","shift","req_type","bill_rate","slots_open","req_open_date","notes"] if c in open_df.columns]
            st.dataframe(open_df[disp_cols].rename(columns={
                "specialty":"Specialty","job_title":"Job Title","discipline":"Discipline",
                "shift":"Shift","req_type":"Type","bill_rate":"Bill Rate",
                "slots_open":"Slots","req_open_date":"Req Opened","notes":"Notes"
            }), use_container_width=True, hide_index=True)

    # Active candidates table
    if not df_cands.empty:
        st.markdown('<div class="section-header" style="margin-top:1rem">Active Candidates</div>', unsafe_allow_html=True)
        active_df = df_cands[df_cands["status"].isin(["Submitted","Clinical Call Scheduled",
                             "Clinical Call Complete","Offered","Accepted"])].copy()
        if active_df.empty:
            st.info("No candidates currently in pipeline.")
        else:
            # Calculate days in pipeline
            active_df["days_in_pipe"] = active_df["date_sent"].apply(
                lambda x: (date.today() - date.fromisoformat(str(x))).days if x and str(x) != "nan" else ""
            )
            disp_c = [c for c in ["candidate_name","source_company","specialty","status",
                                   "date_sent","date_offered","start_date","days_in_pipe","rmchcs_notes"] if c in active_df.columns]
            st.dataframe(active_df[disp_c].rename(columns={
                "candidate_name":"Candidate","source_company":"Source","specialty":"Specialty",
                "status":"Status","date_sent":"Sent","date_offered":"Offered",
                "start_date":"Start","days_in_pipe":"Days in Pipe","rmchcs_notes":"RMCHCS Notes"
            }), use_container_width=True, hide_index=True)

    # Turnaround stats
    if not df_cands.empty and "date_sent" in df_cands.columns and "date_accepted" in df_cands.columns:
        accepted = df_cands[df_cands["date_accepted"].notna() & (df_cands["date_accepted"] != "")].copy()
        if not accepted.empty:
            st.markdown('<div class="section-header" style="margin-top:1rem">Turnaround Times</div>', unsafe_allow_html=True)
            def days_between(a, b):
                try: return (date.fromisoformat(str(b)) - date.fromisoformat(str(a))).days
                except: return None
            accepted["submit_to_accept"] = accepted.apply(lambda r: days_between(r["date_sent"], r["date_accepted"]), axis=1)
            avg_tat = accepted["submit_to_accept"].dropna().mean()
            t1,t2 = st.columns(2)
            t1.metric("Avg Submission to Acceptance", f"{avg_tat:.0f} days" if not pd.isna(avg_tat) else "N/A")
            t2.metric("Total Accepted", len(accepted))


# ══════════════════════════════════════════════════════════════════════════════
# REQUISITIONS
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Requisitions":
    st.markdown("""
    <div class="page-header">
        <div>
            <div class="page-header-title">Requisitions</div>
            <div class="page-header-sub">Open needs and backfill positions</div>
        </div>
    </div>""", unsafe_allow_html=True)

    tab_view, tab_add, tab_import = st.tabs(["All Requisitions", "Add Requisition", "Import from Excel"])

    # ── All Requisitions ─────────────────────────────────────────────────────
    with tab_view:
        st.markdown("<div style='height:0.75rem'></div>", unsafe_allow_html=True)
        if df_reqs.empty:
            st.info("No requisitions yet. Use the Add Requisition tab or Import from Excel.")
        else:
            fc1,fc2,fc3,fc4 = st.columns([2,2,2,2])
            filt_status = fc1.selectbox("Status", ["All"] + REQ_STATUSES, key="req_filt_status")
            filt_disc   = fc2.selectbox("Discipline", ["All"] + DISCIPLINES, key="req_filt_disc")
            filt_type   = fc3.selectbox("Type", ["All"] + REQ_TYPES, key="req_filt_type")
            filt_search = fc4.text_input("Search specialty / notes", key="req_search")

            view_reqs = df_reqs.copy()
            if filt_status != "All": view_reqs = view_reqs[view_reqs["status"]==filt_status]
            if filt_disc   != "All": view_reqs = view_reqs[view_reqs["discipline"]==filt_disc]
            if filt_type   != "All": view_reqs = view_reqs[view_reqs["req_type"]==filt_type]
            if filt_search:
                mask = (view_reqs["specialty"].str.contains(filt_search, case=False, na=False) |
                        view_reqs["notes"].str.contains(filt_search, case=False, na=False))
                view_reqs = view_reqs[mask]

            view_reqs = view_reqs.sort_values("req_open_date", ascending=False).reset_index(drop=True)
            st.markdown(f"**{len(view_reqs)} requisitions** — click any cell to edit, then hit Save All Changes.")

            edit_cols = ["status","discipline","specialty","shift","req_type","bill_rate","slots_open","req_open_date","notes"]
            edit_cols = [c for c in edit_cols if c in view_reqs.columns]

            edited_reqs = st.data_editor(
                view_reqs[edit_cols + ["id"]].rename(columns={
                    "status":"Status","discipline":"Discipline","specialty":"Specialty",
                    "shift":"Shift","req_type":"Type","bill_rate":"Bill Rate",
                    "slots_open":"Slots","req_open_date":"Opened","notes":"Notes","id":"_id"
                }),
                column_config={
                    "Status":    st.column_config.SelectboxColumn("Status",    options=REQ_STATUSES, required=True),
                    "Discipline":st.column_config.SelectboxColumn("Discipline",options=DISCIPLINES,  required=True),
                    "Type":      st.column_config.SelectboxColumn("Type",      options=REQ_TYPES,    required=True),
                    "Bill Rate": st.column_config.NumberColumn("Bill Rate",    min_value=0, step=1, format="$%.0f"),
                    "Slots":     st.column_config.NumberColumn("Slots",        min_value=0, step=1),
                    "Opened":    st.column_config.TextColumn("Opened"),
                    "Notes":     st.column_config.TextColumn("Notes",  width="large"),
                    "Specialty": st.column_config.TextColumn("Specialty"),
                    "Shift":     st.column_config.TextColumn("Shift"),
                    "_id":       st.column_config.Column("_id", disabled=True, width="small"),
                },
                use_container_width=True, hide_index=True,
                height=min(400, 55 + len(view_reqs) * 35),
                key="req_editor"
            )

            sc1,sc2,_ = st.columns([1,1,4])
            with sc1:
                if st.button("Save All Changes", type="primary", use_container_width=True, key="req_save_all"):
                    col_map = {"Status":"status","Discipline":"discipline","Specialty":"specialty",
                               "Shift":"shift","Type":"req_type","Bill Rate":"bill_rate",
                               "Slots":"slots_open","Opened":"req_open_date","Notes":"notes"}
                    saved = 0
                    cascaded = 0
                    for _, row in edited_reqs.iterrows():
                        rec_id  = row["_id"]
                        updates = {col_map[k]: row[k] for k in col_map if k in row}
                        _update_record(REQ_FILE, rec_id, updates)
                        saved += 1
                        # Cascade: any req saved as Closed or Filled floods active candidates to Job Closed
                        if updates.get("status") in ("Closed", "Filled"):
                            cands_data = _load(CANDIDATE_FILE)
                            changed = False
                            for c in cands_data:
                                if c.get("req_id") == rec_id and c.get("status") in ACTIVE_CAND_STATUSES:
                                    c["status"] = "Job Closed"
                                    c["updated_at"] = datetime.now().isoformat()
                                    cascaded += 1
                                    changed = True
                            if changed:
                                with open(CANDIDATE_FILE, "w") as f:
                                    json.dump(cands_data, f, indent=2)
                    msg = f"{saved} requisition(s) saved."
                    if cascaded:
                        msg += f" {cascaded} candidate(s) automatically set to Job Closed."
                    st.success(msg)
                    st.rerun()
            with sc2:
                del_req_labels = [f"{r.get('Specialty','')} | {r.get('Status','')}" for _, r in edited_reqs.iterrows()]
                del_choice = st.selectbox("Delete a req", ["— select to delete —"] + del_req_labels, key="req_del_sel", label_visibility="collapsed")
                if del_choice != "— select to delete —":
                    del_idx = del_req_labels.index(del_choice)
                    del_id  = edited_reqs.iloc[del_idx]["_id"]
                    if st.button("Confirm Delete", type="secondary", key="req_del_btn"):
                        _delete_record(REQ_FILE, del_id)
                        st.success("Deleted.")
                        st.rerun()

            # Candidates linked to any selected req — shown below grid
            if not df_cands.empty and "req_id" in df_cands.columns:
                req_ids_shown = view_reqs["id"].tolist()
                linked = df_cands[df_cands["req_id"].isin(req_ids_shown)]
                if not linked.empty:
                    st.markdown("**Candidates on these requisitions:**")
                    lc = linked[["candidate_name","source_company","specialty","status","date_sent","start_date"]].copy()
                    lc.columns = ["Candidate","Source","Specialty","Status","Date Sent","Start Date"]
                    st.dataframe(lc, use_container_width=True, hide_index=True)

    # ── Add Requisition ──────────────────────────────────────────────────────
    with tab_add:
        st.markdown("<div style='height:0.75rem'></div>", unsafe_allow_html=True)
        c1,c2 = st.columns(2)
        with c1:
            r_specialty  = st.text_input("Specialty", placeholder="e.g. ER RN, Ortho Surgery, Surg Tech")
            r_job_title  = st.text_input("Job Title", placeholder="e.g. OR RN, Nuclear Med Tech")
            r_discipline = st.selectbox("Discipline", DISCIPLINES)
            r_shift      = st.text_input("Shift", placeholder="e.g. 12H Nights, 8H Days")
        with c2:
            r_req_type  = st.selectbox("Type", REQ_TYPES)
            r_bill_rate = st.number_input("Bill Rate ($/hr)", min_value=0.0, value=83.0, step=1.0)
            r_slots     = st.number_input("Slots Open", min_value=1, value=1, step=1)
            r_open_date = st.date_input("Req Open Date", value=date.today())
        r_status = st.selectbox("Status", REQ_STATUSES)
        r_notes  = st.text_area("Notes", placeholder="Client contact, context, hold reason...", height=80)
        rb1, _ = st.columns([1,4])
        with rb1:
            if st.button("Save Requisition", type="primary", use_container_width=True):
                if not r_specialty.strip():
                    st.error("Specialty is required.")
                else:
                    _save_record(REQ_FILE, {
                        "specialty": r_specialty.strip(), "job_title": r_job_title.strip(),
                        "discipline": r_discipline, "shift": r_shift.strip(),
                        "req_type": r_req_type, "bill_rate": r_bill_rate,
                        "slots_open": r_slots, "req_open_date": str(r_open_date),
                        "status": r_status, "notes": r_notes.strip()
                    })
                    st.success(f"Saved: {r_specialty} | {r_req_type} | {r_status}")
                    st.rerun()

    with tab_import:
        st.markdown("<div style='height:0.75rem'></div>", unsafe_allow_html=True)
        st.markdown('<div class="section-header">Import from Excel</div>', unsafe_allow_html=True)
        st.markdown('<div class="section-sub">Upload your RMCHCS candidates and jobs Excel file. Reqs, candidates, and credentialing are all imported in one shot.</div>', unsafe_allow_html=True)

        imp_file = st.file_uploader("Choose Excel file (.xlsx)", type=["xlsx"], label_visibility="collapsed", key="pipeline_import")

        if imp_file:
            with st.spinner("Parsing file..."):
                try:
                    result = parse_pipeline_excel(imp_file)
                    imp_reqs   = result["reqs"]
                    imp_cands  = result["candidates"]
                    imp_warns  = result["warnings"]
                except Exception as e:
                    st.error(f"Could not parse file: {e}")
                    imp_reqs, imp_cands, imp_warns = [], [], []

            for w in imp_warns:
                st.warning(w)

            if not imp_reqs and not imp_cands:
                st.error("Nothing found to import.")
            else:
                ik1,ik2,ik3 = st.columns(3)
                ik1.metric("Requisitions Found",  len(imp_reqs))
                ik2.metric("Candidates Found",    len(imp_cands))
                ik3.metric("With Cred Records",   len([c for c in imp_cands if c.get("cred_company")]))

                st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)

                if imp_reqs:
                    st.markdown("**Requisitions Preview**")
                    req_prev = pd.DataFrame(imp_reqs)[["discipline","specialty","shift","req_type","bill_rate","slots_open","status","notes"]]
                    req_prev.columns = ["Discipline","Specialty","Shift","Type","Bill Rate","Slots","Status","Notes"]
                    st.dataframe(req_prev, use_container_width=True, hide_index=True)

                if imp_cands:
                    st.markdown("**Candidates Preview**")
                    cand_prev = pd.DataFrame(imp_cands)[["candidate_name","source_company","discipline","specialty","status","date_sent","start_date","cred_company","cred_status"]]
                    cand_prev.columns = ["Candidate","Source","Discipline","Specialty","Status","Date Sent","Start Date","Cred Co","Cred Status"]
                    st.dataframe(cand_prev, use_container_width=True, hide_index=True)

                st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)
                st.info("Existing records are NOT overwritten. This adds the imported records alongside anything already in the system.")

                ii1, _ = st.columns([1,4])
                with ii1:
                    if st.button("Import All", type="primary", use_container_width=True):
                        for r in imp_reqs:
                            _save_record(REQ_FILE, r)
                        for c in imp_cands:
                            _save_record(CANDIDATE_FILE, c)
                        st.success(f"Imported {len(imp_reqs)} requisitions and {len(imp_cands)} candidates.")
                        st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# CANDIDATES
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Candidates":
    st.markdown("""
    <div class="page-header">
        <div>
            <div class="page-header-title">Candidates</div>
            <div class="page-header-sub">Submitted candidates and lifecycle tracking</div>
        </div>
    </div>""", unsafe_allow_html=True)

    tab_view, tab_add = st.tabs(["All Candidates", "Add Candidate"])

    # ── All Candidates ───────────────────────────────────────────────────────
    with tab_view:
        st.markdown("<div style='height:0.75rem'></div>", unsafe_allow_html=True)
        if df_cands.empty:
            st.info("No candidates yet. Use the Add Candidate tab or import from Excel on the Requisitions page.")
        else:
            fc1,fc2,fc3,fc4 = st.columns([2,2,2,2])
            filt_cs   = fc1.selectbox("Status", ["All"] + CAND_STATUSES, key="cf_status")
            filt_src  = fc2.selectbox("Source", ["All"] + SOURCE_COS, key="cf_source")
            filt_disc = fc3.selectbox("Discipline", ["All"] + DISCIPLINES, key="cf_disc")
            filt_srch = fc4.text_input("Search name / specialty", key="cf_search")

            vc = df_cands.copy()
            if filt_cs   != "All": vc = vc[vc["status"]==filt_cs]
            if filt_src  != "All": vc = vc[vc["source_company"]==filt_src]
            if filt_disc != "All": vc = vc[vc["discipline"]==filt_disc]
            if filt_srch:
                mask = (vc["candidate_name"].str.contains(filt_srch, case=False, na=False) |
                        vc["specialty"].str.contains(filt_srch, case=False, na=False))
                vc = vc[mask]

            vc = vc.sort_values("date_sent", ascending=False).reset_index(drop=True)

            def _days_in(d):
                try: return int((date.today() - date.fromisoformat(str(d))).days)
                except: return None

            vc["days_in_pipe"] = vc["date_sent"].apply(_days_in)

            st.markdown(f"**{len(vc)} candidates** — click any cell to edit, then hit Save All Changes.")
            st.caption("Credentialing fields are in the rightmost columns. Scroll right to see them.")

            cand_edit_cols = ["status","candidate_name","source_company","discipline","specialty",
                              "date_sent","date_offered","date_accepted","start_date",
                              "days_in_pipe","rmchcs_notes","notes",
                              "cred_company","cred_due_date","cred_status","cred_notes","id"]
            cand_edit_cols = [c for c in cand_edit_cols if c in vc.columns]

            edited_cands = st.data_editor(
                vc[cand_edit_cols].rename(columns={
                    "status":"Status","candidate_name":"Candidate","source_company":"Source",
                    "discipline":"Discipline","specialty":"Specialty",
                    "date_sent":"Sent","date_offered":"Offered","date_accepted":"Accepted",
                    "start_date":"Start","days_in_pipe":"Days","rmchcs_notes":"RMCHCS Notes",
                    "notes":"Internal Notes","cred_company":"Cred Co",
                    "cred_due_date":"Cred Due","cred_status":"Cred Status",
                    "cred_notes":"Cred Notes","id":"_id"
                }),
                column_config={
                    "Status":    st.column_config.SelectboxColumn("Status",    options=CAND_STATUSES, required=True, width="medium"),
                    "Source":    st.column_config.SelectboxColumn("Source",    options=SOURCE_COS,    width="small"),
                    "Discipline":st.column_config.SelectboxColumn("Discipline",options=DISCIPLINES,   width="small"),
                    "Cred Status":st.column_config.SelectboxColumn("Cred Status", options=[""]+CRED_STATUSES, width="small"),
                    "Candidate": st.column_config.TextColumn("Candidate",  width="medium"),
                    "Specialty": st.column_config.TextColumn("Specialty",  width="medium"),
                    "Sent":      st.column_config.TextColumn("Sent",       width="small"),
                    "Offered":   st.column_config.TextColumn("Offered",    width="small"),
                    "Accepted":  st.column_config.TextColumn("Accepted",   width="small"),
                    "Start":     st.column_config.TextColumn("Start",      width="small"),
                    "Days":      st.column_config.NumberColumn("Days",     disabled=True, width="small"),
                    "RMCHCS Notes":   st.column_config.TextColumn("RMCHCS Notes",   width="large"),
                    "Internal Notes": st.column_config.TextColumn("Internal Notes", width="large"),
                    "Cred Co":   st.column_config.TextColumn("Cred Co",   width="small"),
                    "Cred Due":  st.column_config.TextColumn("Cred Due",  width="small"),
                    "Cred Notes":st.column_config.TextColumn("Cred Notes",width="medium"),
                    "_id":       st.column_config.Column("_id", disabled=True, width="small"),
                },
                use_container_width=True, hide_index=True,
                height=min(500, 55 + len(vc) * 35),
                key="cand_editor"
            )

            sc1,sc2,_ = st.columns([1,1,4])
            with sc1:
                if st.button("Save All Changes", type="primary", use_container_width=True, key="cand_save_all"):
                    col_map = {
                        "Status":"status","Source":"source_company","Discipline":"discipline",
                        "Specialty":"specialty","Sent":"date_sent","Offered":"date_offered",
                        "Accepted":"date_accepted","Start":"start_date",
                        "RMCHCS Notes":"rmchcs_notes","Internal Notes":"notes",
                        "Cred Co":"cred_company","Cred Due":"cred_due_date",
                        "Cred Status":"cred_status","Cred Notes":"cred_notes"
                    }
                    saved = 0
                    for _, row in edited_cands.iterrows():
                        rec_id  = row["_id"]
                        updates = {col_map[k]: row[k] for k in col_map if k in row}
                        _update_record(CANDIDATE_FILE, rec_id, updates)
                        saved += 1
                    st.success(f"{saved} candidate(s) saved.")
                    st.rerun()
            with sc2:
                del_cand_labels = [f"{r.get('Candidate','')} | {r.get('Status','')}" for _, r in edited_cands.iterrows()]
                del_choice = st.selectbox("Delete", ["— select to delete —"] + del_cand_labels, key="cand_del_sel", label_visibility="collapsed")
                if del_choice != "— select to delete —":
                    del_idx = del_cand_labels.index(del_choice)
                    del_id  = edited_cands.iloc[del_idx]["_id"]
                    if st.button("Confirm Delete", type="secondary", key="cand_del_btn"):
                        _delete_record(CANDIDATE_FILE, del_id)
                        st.success("Deleted.")
                        st.rerun()

    # ── Add Candidate ────────────────────────────────────────────────────────
    with tab_add:
        st.markdown("<div style='height:0.75rem'></div>", unsafe_allow_html=True)
        ca1,ca2 = st.columns(2)
        with ca1:
            c_name      = st.text_input("Candidate Name", placeholder="First Last")
            c_source    = st.selectbox("Source Company", SOURCE_COS)
            c_disc      = st.selectbox("Discipline", DISCIPLINES, key="c_disc")
            c_specialty = st.text_input("Specialty", placeholder="e.g. ER RN, Ortho Surgery")
        with ca2:
            req_options = ["None"]
            if not df_reqs.empty:
                req_options += [f"{r['specialty']} | {r.get('discipline','')} | {r.get('status','')} [{r['id'][:8]}]"
                                for _, r in df_reqs.iterrows()]
            linked_req_label = st.selectbox("Link to Requisition (optional)", req_options)
            linked_req_id = None
            if linked_req_label != "None":
                short_id = linked_req_label.split("[")[-1].rstrip("]")
                matches = df_reqs[df_reqs["id"].str.startswith(short_id)]
                if not matches.empty:
                    linked_req_id = matches.iloc[0]["id"]
            c_status = st.selectbox("Current Status", CAND_STATUSES)
            c_shift  = st.text_input("Shift", placeholder="e.g. 12H Nights")

        st.markdown("---")
        st.markdown("**Timeline**")
        cb1,cb2,cb3 = st.columns(3)
        with cb1:
            c_date_sent = st.date_input("Date Sent to Client", value=date.today())
            c_date_clin = st.date_input("Clinical Call Date", value=None)
        with cb2:
            c_date_offered  = st.date_input("Date Offered", value=None)
            c_date_accepted = st.date_input("Date Accepted", value=None)
        with cb3:
            c_start_date = st.date_input("Start Date", value=None)
            if c_date_sent and c_date_accepted:
                st.metric("Submission → Acceptance", f"{(c_date_accepted - c_date_sent).days} days")

        st.markdown("---")
        cn1,cn2 = st.columns(2)
        c_notes        = cn1.text_area("Internal Notes", height=80)
        c_rmchcs_notes = cn2.text_area("RMCHCS Notes", height=80, placeholder="Backfill for X, Curry calling, etc.")

        save_b, _ = st.columns([1,4])
        with save_b:
            if st.button("Save Candidate", type="primary", use_container_width=True):
                if not c_name.strip():
                    st.error("Candidate name is required.")
                else:
                    _save_record(CANDIDATE_FILE, {
                        "req_id": linked_req_id,
                        "candidate_name": c_name.strip(), "source_company": c_source,
                        "discipline": c_disc, "specialty": c_specialty.strip(),
                        "status": c_status,
                        "date_sent":          str(c_date_sent)     if c_date_sent     else "",
                        "date_clinical_call": str(c_date_clin)     if c_date_clin     else "",
                        "date_offered":       str(c_date_offered)  if c_date_offered  else "",
                        "date_accepted":      str(c_date_accepted) if c_date_accepted else "",
                        "start_date":         str(c_start_date)    if c_start_date    else "",
                        "notes": c_notes.strip(), "rmchcs_notes": c_rmchcs_notes.strip(),
                        "cred_company":"","cred_due_date":"",
                        "cred_status":"","cred_nm_fingerprint":False,"cred_notes":""
                    })
                    st.success(f"Saved: {c_name} | {c_specialty} | {c_status}")
                    st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# CREDENTIALING VIEW
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Credentialing":
    st.markdown("""
    <div class="page-header">
        <div>
            <div class="page-header-title">Credentialing</div>
            <div class="page-header-sub">Status across all active candidates</div>
        </div>
    </div>""", unsafe_allow_html=True)

    if df_cands.empty:
        st.info("No candidates yet.")
        st.stop()

    cred_df = df_cands[df_cands["cred_company"].notna() & (df_cands["cred_company"] != "")].copy()               if "cred_company" in df_cands.columns else pd.DataFrame()

    if cred_df.empty:
        st.info("No credentialing records yet. Add credentialing details from the Candidates page.")
        st.stop()

    today_str = date.today().isoformat()

    # Flag overdue
    def cred_flag(row):
        if row.get("cred_status") in ["Clear","Cancelled"]: return "✅ Clear"
        if row.get("cred_due_date") and str(row.get("cred_due_date","")) <= today_str: return "🔴 Overdue"
        if row.get("cred_due_date"): return "🟡 Pending"
        return "⚪ No Date"

    cred_df["flag"] = cred_df.apply(cred_flag, axis=1)

    # Summary
    ck1,ck2,ck3,ck4 = st.columns(4)
    ck1.metric("Total Credentialing", len(cred_df))
    ck2.metric("Clear",    len(cred_df[cred_df["flag"]=="✅ Clear"]))
    ck3.metric("Pending",  len(cred_df[cred_df["flag"]=="🟡 Pending"]))
    ck4.metric("Overdue",  len(cred_df[cred_df["flag"]=="🔴 Overdue"]))

    st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)

    filt_cred = st.selectbox("Filter by Status", ["All","🔴 Overdue","🟡 Pending","✅ Clear","⚪ No Date"])
    view_cred = cred_df if filt_cred=="All" else cred_df[cred_df["flag"]==filt_cred]

    disp_cred = view_cred[["flag","candidate_name","specialty","cred_company",
                             "cred_due_date","cred_status","cred_nm_fingerprint","cred_notes"]].copy()
    disp_cred["cred_nm_fingerprint"] = disp_cred["cred_nm_fingerprint"].apply(lambda x: "Yes" if x else "No")
    disp_cred.columns = ["Flag","Candidate","Specialty","Cred Company",
                          "Due Date","Status","NM Fingerprint","Notes"]
    st.dataframe(disp_cred.sort_values("Due Date"), use_container_width=True, hide_index=True)

# Separator page (divider in sidebar)

# ══════════════════════════════════════════════════════════════════════════════
# PROGRAM ROI  — DiagnOS-powered
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Program ROI":
    # ── Revenue benchmarks pulled from DiagnOS data ───────────────────────────
    SPECIALTY_REVENUE = {
        "Allergy/Immunology":          (280.0,   "encounters"),
        "Anesthesia (MD)":             (500.0,   "cases"),
        "Cardiology (Non-Invasive)":   (4611.0,  "cases"),
        "Cardiothoracic Surgery":      (8500.0,  "cases"),
        "Cardiovascular Surgery":      (8000.0,  "cases"),
        "Critical Care":               (500.0,   "patient-days"),
        "Dermatology":                 (380.0,   "encounters"),
        "Emergency Medicine":          (620.0,   "visits"),
        "Endocrinology":               (185.0,   "encounters"),
        "ENT (Otolaryngology)":        (1400.0,  "cases"),
        "Family Practice/Primary Care":(185.0,   "visits"),
        "Gastroenterology (GI)":       (1362.0,  "cases"),
        "GI Motility Specialist":      (320.0,   "encounters"),
        "General Surgery":             (3200.0,  "cases"),
        "Hematology Oncology":         (1850.0,  "encounters"),
        "Hospitalist":                 (880.0,   "visits"),
        "Infectious Disease":          (200.0,   "encounters"),
        "Internal Medicine":           (175.0,   "encounters"),
        "Interventional Cardiology":   (8200.0,  "cases"),
        "Medical Oncology":            (1500.0,  "encounters"),
        "Nephrology":                  (220.0,   "encounters"),
        "Neurology":                   (350.0,   "encounters"),
        "Neurosurgery":                (6500.0,  "cases"),
        "OB/GYN":                      (250.0,   "visits"),
        "Orthopedic Surgery":          (6419.0,  "cases"),
        "Orthopedic - Hand Surgery":   (5800.0,  "cases"),
        "Orthopedic - Trauma Surgery": (6800.0,  "cases"),
        "Pathology":                   (180.0,   "cases"),
        "Pediatrics":                  (140.0,   "encounters"),
        "Physical Medicine":           (280.0,   "encounters"),
        "Psychiatry":                  (275.0,   "visits"),
        "Pulmonology":                 (280.0,   "encounters"),
        "Radiation Oncology":          (1800.0,  "cases"),
        "Radiology":                   (150.0,   "reads"),
        "Radiology - Interventional":  (1200.0,  "procedures"),
        "Rheumatology":                (200.0,   "encounters"),
        "Trauma Surgery":              (7500.0,  "cases"),
        "Urgent Care":                 (210.0,   "visits"),
        "Urology":                     (1887.0,  "cases"),
        "Vascular Surgery":            (4500.0,  "cases"),
        # APP
        "APP - Cardiology":            (380.0,   "visits"),
        "APP - Critical Care":         (320.0,   "patient-days"),
        "APP - Emergency Medicine":    (440.0,   "visits"),
        "APP - General Surgery":       (900.0,   "cases"),
        "APP - Hospitalist":           (680.0,   "visits"),
        "APP - Neurology":             (240.0,   "encounters"),
        "APP - Oncology":              (750.0,   "encounters"),
        "APP - Orthopedic Surgery":    (900.0,   "cases"),
        "APP - Psychiatry (PMHNP)":    (220.0,   "visits"),
        "CRNA":                        (350.0,   "cases"),
        "Other / Custom":              (0.0,     "encounters"),
    }

    # ── ROI volume config (persisted in session state) ─────────────────────────
    ROI_FILE = "roi_config.json"
    def _load_roi():
        if os.path.exists(ROI_FILE):
            with open(ROI_FILE) as f:
                return json.load(f)
        return {}
    def _save_roi(data):
        with open(ROI_FILE, "w") as f:
            json.dump(data, f, indent=2)

    roi_config = _load_roi()

    st.markdown("""
    <div class="page-header">
        <div>
            <div class="page-header-title">Program ROI</div>
            <div class="page-header-sub">Real revenue vs actual labor cost — powered by DiagnOS benchmarks</div>
        </div>
    </div>""", unsafe_allow_html=True)

    # ── Pull placed providers from candidates + their spend ───────────────────
    placed = df_cands[df_cands["status"] == "Placed"].copy() if not df_cands.empty else pd.DataFrame()

    if placed.empty and df.empty:
        st.info("No placed candidates or logged spend yet. Place a candidate and log spend to see ROI.")
    else:
        # ── Step 1: MD/APP Revenue Providers config ──────────────────────────────
        # MD/APP disciplines drive revenue — Nursing/Allied are cost-only
        MD_APP_DISCIPLINES = {"Physician", "APP", "Locums"}
        NURSING_ALLIED_DISCIPLINES = {"Nursing", "Allied Health"}

        st.markdown("#### MD & APP Revenue Configuration")
        st.caption("Revenue providers only. Specialty and reimbursement rate pre-fill from DiagnOS benchmarks — the rate field updates when you change the specialty dropdown. Override with RMCHCS actual rates as needed.")

        provider_rows = []

        # Filter placed to MD/APP only
        if not placed.empty:
            revenue_placed = placed[
                placed.get("discipline", pd.Series(dtype=str)).isin(MD_APP_DISCIPLINES) |
                placed["discipline"].isna()
            ].copy() if "discipline" in placed.columns else placed.copy()
            # Exclude clearly allied/nursing names if discipline missing
            revenue_placed = revenue_placed[
                ~revenue_placed.get("discipline", pd.Series(dtype=str)).isin(NURSING_ALLIED_DISCIPLINES)
            ] if "discipline" in revenue_placed.columns else revenue_placed
            revenue_provider_names = revenue_placed["candidate_name"].dropna().unique().tolist()
            cost_only_placed = placed[
                placed.get("discipline", pd.Series(dtype=str)).isin(NURSING_ALLIED_DISCIPLINES)
            ].copy() if "discipline" in placed.columns else pd.DataFrame()
        else:
            revenue_provider_names = []
            cost_only_placed = pd.DataFrame()

        specialty_options = sorted(SPECIALTY_REVENUE.keys())

        if not revenue_provider_names:
            st.info("No placed MD/APP providers yet. Once candidates with Physician, APP, or Locums discipline are placed, they appear here.")
        else:
            cols_hdr = st.columns([3,2,2,2,1])
            cols_hdr[0].markdown("**Provider**")
            cols_hdr[1].markdown("**Specialty**")
            cols_hdr[2].markdown("**Rate/Unit ($)**")
            cols_hdr[3].markdown("**Units/Week**")
            cols_hdr[4].markdown("**Unit**")

            st.markdown("<hr style='margin:4px 0 8px 0; border-color:#e2e8f0'>", unsafe_allow_html=True)

            for pname in revenue_provider_names:
                cfg = roi_config.get(pname, {})

                # Get specialty from candidate record
                if not revenue_placed.empty:
                    match = revenue_placed[revenue_placed["candidate_name"] == pname]
                    cand_spec = match.iloc[0].get("specialty", "") if not match.empty else ""
                else:
                    cand_spec = ""

                # Resolve specialty: saved config → candidate field → "Other / Custom"
                saved_spec = cfg.get("specialty", "")
                if saved_spec and saved_spec in SPECIALTY_REVENUE:
                    resolved_spec = saved_spec
                elif cand_spec and cand_spec in SPECIALTY_REVENUE:
                    resolved_spec = cand_spec
                else:
                    resolved_spec = "Other / Custom"

                c1, c2, c3, c4, c5 = st.columns([3,2,2,2,1])
                c1.markdown(f"**{pname}**")

                spec_idx = specialty_options.index(resolved_spec) if resolved_spec in specialty_options else 0
                sel_spec = c2.selectbox("Specialty", specialty_options,
                                        index=spec_idx,
                                        key=f"roi_spec_{pname}",
                                        label_visibility="collapsed")

                # Rate: ALWAYS drive from current dropdown selection first,
                # then fall back to saved override only if specialty matches saved
                benchmark_rate, unit_val = SPECIALTY_REVENUE.get(sel_spec, (0.0, "encounters"))
                saved_rate = cfg.get("rate", None)
                saved_spec_for_rate = cfg.get("specialty", "")
                # Use saved rate only if it was saved against THIS same specialty
                if saved_rate is not None and saved_spec_for_rate == sel_spec:
                    init_rate = float(saved_rate)
                else:
                    init_rate = float(benchmark_rate)

                override_rate = c3.number_input("Rate", value=init_rate,
                                                min_value=0.0, step=10.0,
                                                key=f"roi_rate_{pname}_{sel_spec}",
                                                label_visibility="collapsed")
                vol_per_wk = c4.number_input("Vol/Wk",
                                              value=float(cfg.get("volume_per_week", 20.0)),
                                              min_value=0.0, step=1.0,
                                              key=f"roi_vol_{pname}",
                                              label_visibility="collapsed")
                c5.markdown(
                    f"<div style='padding-top:8px;font-size:0.78rem;color:#64748b'>{unit_val}</div>",
                    unsafe_allow_html=True)

                provider_rows.append({
                    "provider":        pname,
                    "specialty":       sel_spec,
                    "rate":            override_rate,
                    "volume_per_week": vol_per_wk,
                    "unit":            unit_val,
                })

            sv1, sv2, _ = st.columns([1,1,4])
            if sv1.button("Save Configuration", type="primary", use_container_width=True):
                for row in provider_rows:
                    roi_config[row["provider"]] = {
                        "specialty":       row["specialty"],
                        "rate":            row["rate"],
                        "volume_per_week": row["volume_per_week"],
                    }
                _save_roi(roi_config)
                st.success("Configuration saved.")
                st.rerun()

        # ── Cost-only section: Nursing / Allied ───────────────────────────────
        if not cost_only_placed.empty:
            st.markdown("---")
            st.markdown("#### Nursing & Allied Health — Cost Tracking")
            st.caption("These providers are cost lines only. No revenue attribution.")
            cost_names = cost_only_placed["candidate_name"].dropna().unique().tolist()
            cost_data = []
            for cname in cost_names:
                disc = cost_only_placed[cost_only_placed["candidate_name"]==cname]["discipline"].values
                disc_val = disc[0] if len(disc) > 0 else "Allied Health"
                cost_data.append({"Provider": cname, "Discipline": disc_val})
            st.dataframe(pd.DataFrame(cost_data), use_container_width=True, hide_index=True)

            st.markdown("---")

            # ── Step 2: Date range for calculation ────────────────────────────
            st.markdown("#### Analysis Period")
            dc1, dc2 = st.columns(2)
            try:
                min_date = date.fromisoformat(str(df["week_ending"].min())) if not df.empty else date.today() - timedelta(days=90)
                max_date = date.fromisoformat(str(df["week_ending"].max())) if not df.empty else date.today()
            except Exception:
                min_date = date.today() - timedelta(days=90)
                max_date = date.today()

            start_date = dc1.date_input("From", value=min_date, key="roi_start")
            end_date   = dc2.date_input("To",   value=max_date, key="roi_end")

            # ── Step 3: Compute ROI ───────────────────────────────────────────
            if not df.empty:
                spend_filtered = df.copy()
                try:
                    spend_filtered["_we"] = pd.to_datetime(spend_filtered["week_ending"]).dt.date
                    spend_filtered = spend_filtered[
                        (spend_filtered["_we"] >= start_date) &
                        (spend_filtered["_we"] <= end_date)
                    ]
                except Exception:
                    pass
            else:
                spend_filtered = pd.DataFrame()

            # Number of weeks in period
            period_days  = max(1, (end_date - start_date).days)
            period_weeks = period_days / 7.0

            results = []
            for row in provider_rows:
                pname    = row["provider"]
                rate     = row["rate"]
                vol_wk   = row["volume_per_week"]
                unit     = row["unit"]

                # Actual spend for this provider
                if not spend_filtered.empty and "provider_name" in spend_filtered.columns:
                    provider_spend = spend_filtered[
                        spend_filtered["provider_name"].str.lower().str.contains(
                            pname.split()[0].lower(), na=False)
                    ]["total_spend"].sum()
                else:
                    provider_spend = 0.0

                # Estimated revenue
                est_revenue  = vol_wk * rate * period_weeks
                net_contrib  = est_revenue - provider_spend
                roi_pct      = ((net_contrib / provider_spend) * 100) if provider_spend > 0 else 0.0

                results.append({
                    "Provider":          pname,
                    "Specialty":         row["specialty"],
                    "Labor Cost":        provider_spend,
                    "Est. Revenue":      est_revenue,
                    "Net Contribution":  net_contrib,
                    "ROI %":             roi_pct,
                    "Units/Wk":          vol_wk,
                    "Rate/Unit":         rate,
                    "Unit":              unit,
                })

            # ── Cost-only rows: nursing/allied spend from spend log ──────────────
            cost_only_results = []
            # Collect cost-only provider names (placed nursing/allied)
            cost_only_names = []
            if not cost_only_placed.empty and "candidate_name" in cost_only_placed.columns:
                cost_only_names = cost_only_placed["candidate_name"].dropna().unique().tolist()

            for cname in cost_only_names:
                if not spend_filtered.empty and "provider_name" in spend_filtered.columns:
                    c_spend = spend_filtered[
                        spend_filtered["provider_name"].str.lower().str.contains(
                            cname.split()[0].lower(), na=False)
                    ]["total_spend"].sum()
                else:
                    c_spend = 0.0
                disc = ""
                if not cost_only_placed.empty and "discipline" in cost_only_placed.columns:
                    disc_vals = cost_only_placed[cost_only_placed["candidate_name"]==cname]["discipline"].values
                    disc = disc_vals[0] if len(disc_vals) > 0 else "Allied Health"
                cost_only_results.append({
                    "provider": cname,
                    "discipline": disc,
                    "labor_cost": c_spend,
                })

            # Unmatched spend — any spend log entries not matched to a known provider
            matched_names = [r["Provider"] for r in results] + cost_only_names
            if not spend_filtered.empty and "provider_name" in spend_filtered.columns:
                def _is_matched(pn):
                    pn_lower = str(pn).lower()
                    for mn in matched_names:
                        if mn.split()[0].lower() in pn_lower:
                            return True
                    return False
                unmatched_spend = spend_filtered[
                    ~spend_filtered["provider_name"].apply(_is_matched)
                ]["total_spend"].sum()
            else:
                unmatched_spend = 0.0

            # True total cost = MD/APP matched + cost-only + unmatched
            md_app_cost      = sum(r["Labor Cost"] for r in results)
            cost_only_total  = sum(r["labor_cost"] for r in cost_only_results)
            true_total_cost  = md_app_cost + cost_only_total + unmatched_spend

            if results:
                st.markdown("---")
                st.markdown("#### Program ROI Dashboard")

                total_rev  = sum(r["Est. Revenue"]     for r in results)
                total_net  = total_rev - true_total_cost
                prog_roi   = ((total_net / true_total_cost) * 100) if true_total_cost > 0 else 0.0

                # KPI cards
                k1, k2, k3, k4 = st.columns(4)
                def _kpi(col, label, val, fmt="$", delta=None):
                    if fmt == "$":
                        disp = f"${val:,.0f}"
                    elif fmt == "%":
                        disp = f"{val:.1f}%"
                    else:
                        disp = str(val)
                    delta_html = ""
                    if delta is not None:
                        clr = "#10b981" if delta >= 0 else "#ef4444"
                        sign = "+" if delta >= 0 else ""
                        delta_html = f'<div style="font-size:0.78rem;color:{clr};font-weight:600">{sign}{delta:.1f}%</div>'
                    col.markdown(
                        f'<div class="kpi-card">'
                        f'<div class="kpi-label">{label}</div>'
                        f'<div class="kpi-value">{disp}</div>'
                        f'{delta_html}'
                        f'</div>', unsafe_allow_html=True)

                _kpi(k1, "True Total Cost",    true_total_cost)
                _kpi(k2, "Estimated Revenue",  total_rev)
                _kpi(k3, "Net Contribution",   total_net,  delta=prog_roi)
                _kpi(k4, "Program ROI",        prog_roi, fmt="%")

                st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)

                # ── Provider breakdown table ──────────────────────────────────
                st.markdown("**Revenue Providers (MD / APP)**")
                df_results = pd.DataFrame(results)
                disp_cols  = ["Provider","Specialty","Labor Cost","Est. Revenue","Net Contribution","ROI %","Units/Wk","Rate/Unit","Unit"]
                df_disp    = df_results[disp_cols].copy()
                df_disp["Labor Cost"]       = df_disp["Labor Cost"].apply(lambda x: f"${x:,.0f}")
                df_disp["Est. Revenue"]     = df_disp["Est. Revenue"].apply(lambda x: f"${x:,.0f}")
                df_disp["Net Contribution"] = df_disp["Net Contribution"].apply(lambda x: f"${x:,.0f}")
                df_disp["ROI %"]            = df_disp["ROI %"].apply(lambda x: f"{x:.1f}%")
                df_disp["Rate/Unit"]        = df_disp["Rate/Unit"].apply(lambda x: f"${x:,.0f}")
                st.dataframe(df_disp, use_container_width=True, hide_index=True)

                # ── Cost-only providers ───────────────────────────────────────
                if cost_only_results or unmatched_spend > 0:
                    st.markdown("**Cost-Only Providers (Nursing / Allied Health)**")
                    cost_rows = []
                    for cr in cost_only_results:
                        cost_rows.append({
                            "Provider":   cr["provider"],
                            "Discipline": cr["discipline"],
                            "Labor Cost": f"${cr['labor_cost']:,.0f}",
                            "Revenue":    "—",
                            "Note":       "Cost line only — no revenue attribution",
                        })
                    if unmatched_spend > 0:
                        cost_rows.append({
                            "Provider":   "Other / Unmatched",
                            "Discipline": "—",
                            "Labor Cost": f"${unmatched_spend:,.0f}",
                            "Revenue":    "—",
                            "Note":       "Logged spend not matched to a placed provider",
                        })
                    st.dataframe(pd.DataFrame(cost_rows), use_container_width=True, hide_index=True)

                st.markdown("---")
                st.markdown("#### Program Financials")

                ch1, ch2 = st.columns(2)

                # Left: program-level total cost vs total revenue
                with ch1:
                    st.markdown("**Program Total: Cost vs Revenue**")
                    fig_prog = go.Figure()
                    fig_prog.add_trace(go.Bar(
                        x=["True Total Cost", "Estimated Revenue"],
                        y=[true_total_cost, total_rev],
                        marker_color=["#ef4444", "#10b981"],
                        text=[f"${true_total_cost:,.0f}", f"${total_rev:,.0f}"],
                        textposition="outside",
                        showlegend=False,
                    ))
                    fig_prog.update_layout(
                        height=320,
                        margin=dict(l=10,r=10,t=10,b=10),
                        plot_bgcolor="#f8fafc", paper_bgcolor="#f8fafc",
                        yaxis=dict(tickprefix="$", gridcolor="#e2e8f0", showgrid=True),
                        xaxis=dict(gridcolor="#e2e8f0"),
                    )
                    st.plotly_chart(fig_prog, use_container_width=True)

                # Right: per-provider cost vs revenue (MD/APP only)
                with ch2:
                    st.markdown("**MD / APP: Cost vs Revenue by Provider**")
                    if results:
                        fig_prov = go.Figure()
                        providers_list = [r["Provider"] for r in results]
                        md_costs = [r["Labor Cost"]   for r in results]
                        revs     = [r["Est. Revenue"] for r in results]
                        fig_prov.add_trace(go.Bar(
                            name="Labor Cost", x=providers_list, y=md_costs,
                            marker_color="#ef4444", opacity=0.85
                        ))
                        fig_prov.add_trace(go.Bar(
                            name="Est. Revenue", x=providers_list, y=revs,
                            marker_color="#10b981", opacity=0.85
                        ))
                        fig_prov.update_layout(
                            barmode="group",
                            height=320,
                            margin=dict(l=10,r=10,t=10,b=10),
                            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                            plot_bgcolor="#f8fafc", paper_bgcolor="#f8fafc",
                            yaxis=dict(tickprefix="$", gridcolor="#e2e8f0"),
                            xaxis=dict(gridcolor="#e2e8f0"),
                        )
                        st.plotly_chart(fig_prov, use_container_width=True)
                    else:
                        st.info("No MD/APP providers configured yet.")

                # Methodology note
                # ── Export buttons ────────────────────────────────────────────────
                st.markdown("---")
                st.markdown("#### Export Reports")
                ex1, ex2, _ = st.columns([1,1,3])

                # ── Build exec brief PDF ──────────────────────────────────────────
                def build_roi_exec_brief(total_cost, total_rev, total_net, prog_roi,
                                          results, cost_only_results, unmatched_spend,
                                          start_date, end_date, period_weeks):
                    import io as _io
                    from reportlab.lib.pagesizes import letter
                    from reportlab.lib import colors
                    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                    from reportlab.lib.units import inch
                    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
                    from datetime import date as _date

                    NAVY  = colors.HexColor("#283652")
                    RED   = colors.HexColor("#DD4F20")
                    WHITE = colors.white
                    LGRAY = colors.HexColor("#F1ECE9")
                    MGRAY = colors.HexColor("#B3B5B0")
                    GREEN = colors.HexColor("#065F46")
                    MUTED = colors.HexColor("#6B7280")
                    GREENBG = colors.HexColor("#ECFDF5")
                    W = 7.0 * inch

                    def fmt(v):  return f"${v:,.0f}"
                    def abbrev(v):
                        if abs(v) >= 1_000_000: return f"${v/1_000_000:.1f}M"
                        if abs(v) >= 1_000:     return f"${v/1_000:.0f}K"
                        return fmt(v)

                    def on_page(canvas, doc):
                        canvas.saveState()
                        pw, ph = letter
                        m = 0.65 * inch
                        canvas.setFillColor(NAVY)
                        canvas.rect(0, ph - 0.75*inch, pw, 0.75*inch, fill=1, stroke=0)
                        canvas.setFillColor(RED)
                        canvas.rect(0, 0, 0.10*inch, ph, fill=1, stroke=0)
                        canvas.setFont("Helvetica-Bold", 10)
                        canvas.setFillColor(WHITE)
                        canvas.drawRightString(pw - m, ph - 0.30*inch, "RMCHCS ITO Program  |  Executive Brief")
                        canvas.setFont("Helvetica", 8)
                        canvas.setFillColor(colors.HexColor("#A8C4D4"))
                        canvas.drawRightString(pw - m, ph - 0.48*inch,
                            f"Ingenovis Health  |  {start_date.strftime('%b %d, %Y')} – {end_date.strftime('%b %d, %Y')}  |  {_date.today().strftime('%B %d, %Y')}")
                        canvas.setStrokeColor(RED)
                        canvas.setLineWidth(1.0)
                        canvas.line(m, 0.58*inch, pw - m, 0.58*inch)
                        canvas.setFont("Helvetica", 7)
                        canvas.setFillColor(MUTED)
                        canvas.drawString(m, 0.42*inch, "Confidential  |  Ingenovis Health ITO")
                        canvas.drawRightString(pw - m, 0.42*inch, "Revenue figures are estimated benchmarks — validate with RMCHCS actual payer data")
                        canvas.restoreState()

                    sty = getSampleStyleSheet()
                    def ps(name, fn="Helvetica", fs=9, ld=13, tc=None, **kw):
                        return ParagraphStyle(name, parent=sty["Normal"], fontName=fn,
                                              fontSize=fs, leading=ld,
                                              textColor=tc if tc is not None else colors.HexColor("#1A1A2E"), **kw)
                    def hr(c=NAVY, t=0.75): return HRFlowable(width="100%", thickness=t, color=c, spaceAfter=5, spaceBefore=3)
                    def sec(text): return Paragraph(f"<b>{text.upper()}</b>",
                                    ps("s", fn="Helvetica-Bold", fs=8, tc=RED, ld=10, spaceBefore=5, spaceAfter=2))

                    roi_c  = GREEN if prog_roi >= 0 else RED
                    roi_bg = GREENBG if prog_roi >= 0 else colors.HexColor("#FDEDEC")

                    # Snapshot cards
                    def stat_row(cards):
                        n, cw = len(cards), W / len(cards)
                        hr_ = [Paragraph(f"<b>{lbl}</b>", ps(f"h{i}", fn="Helvetica-Bold", fs=8, tc=WHITE, alignment=1))
                               for i,(v,lbl,ac) in enumerate(cards)]
                        vr  = [Paragraph(f"<b>{v}</b>", ps(f"v{i}", fn="Helvetica-Bold", fs=18, tc=ac, alignment=1, ld=22))
                               for i,(v,lbl,ac) in enumerate(cards)]
                        th = Table([hr_], colWidths=[cw]*n)
                        th.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),NAVY),
                            ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
                            ("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),
                            ("LINEAFTER",(0,0),(-2,-1),0.5,colors.HexColor("#415B80"))]))
                        tb = Table([vr], colWidths=[cw]*n)
                        tb.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),LGRAY),
                            ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
                            ("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),
                            ("LINEAFTER",(0,0),(-2,-1),0.5,MGRAY),
                            ("LINEBEFORE",(0,0),(0,-1),0.5,MGRAY),
                            ("LINEAFTER",(-1,0),(-1,-1),0.5,MGRAY),
                            ("LINEBELOW",(0,-1),(-1,-1),0.5,MGRAY)]))
                        return [th, tb]

                    story = []

                    # Title
                    story.append(Paragraph("RMCHCS ITO Program", ps("t", fn="Helvetica-Bold", fs=18, tc=NAVY, ld=22, spaceBefore=4)))
                    story.append(Paragraph("Program ROI — Executive Brief",
                        ps("sub", fs=11, tc=colors.HexColor("#415B80"), ld=14, spaceAfter=4)))
                    story.append(Paragraph(
                        f"Analysis period: {start_date.strftime('%B %d, %Y')} – {end_date.strftime('%B %d, %Y')}  ({period_weeks:.1f} weeks)  |  Ingenovis Health ITO",
                        ps("meta", fs=8, tc=MUTED, spaceAfter=8)))
                    story.append(HRFlowable(width="100%", thickness=2.0, color=RED, spaceAfter=10))

                    # Snapshot cards
                    snap = [
                        (abbrev(total_cost),  "True Total Cost",     RED),
                        (abbrev(total_rev),   "Estimated Revenue",   NAVY),
                        (abbrev(total_net),   "Net Contribution",    GREEN if total_net >= 0 else RED),
                        (f"{prog_roi:.1f}%",  "Program ROI",         GREEN if prog_roi >= 0 else RED),
                    ]
                    for item in stat_row(snap): story.append(item)
                    story.append(Spacer(1, 10))

                    # ROI headline
                    story.append(Table([[Paragraph(
                        f"<b>The RMCHCS ITO program generated an estimated {abbrev(total_net)} net contribution "
                        f"on {abbrev(total_cost)} in total labor cost — a {prog_roi:.1f}% program ROI.</b>",
                        ps("roi", tc=roi_c, ld=13))]],
                        colWidths=[W],
                        style=TableStyle([("BACKGROUND",(0,0),(-1,-1),roi_bg),
                            ("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8),
                            ("LEFTPADDING",(0,0),(-1,-1),10),
                            ("LINEABOVE",(0,0),(-1,0),2.0,roi_c)])))
                    story.append(Spacer(1, 10))

                    # Revenue provider table
                    story.append(sec("Revenue Providers (MD / APP)"))
                    story.append(hr())
                    rev_rows = [["Provider","Specialty","Labor Cost","Est. Revenue","Net Contribution","ROI %"]]
                    for r in results:
                        roi_str = f"{r['ROI %']:.1f}%" if isinstance(r['ROI %'], (int,float)) else str(r['ROI %'])
                        rev_rows.append([r["Provider"], r["Specialty"],
                                          fmt(r["Labor Cost"]), fmt(r["Est. Revenue"]),
                                          fmt(r["Net Contribution"]), roi_str])
                    t_rev = Table(rev_rows, colWidths=[1.4*inch,1.5*inch,0.95*inch,0.95*inch,1.1*inch,0.7*inch])
                    t_rev.setStyle(TableStyle([
                        ("BACKGROUND",(0,0),(-1,0),NAVY),("TEXTCOLOR",(0,0),(-1,0),WHITE),
                        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,0),8),
                        ("ROWBACKGROUNDS",(0,1),(-1,-1),[WHITE,LGRAY]),
                        ("FONTNAME",(0,1),(-1,-1),"Helvetica"),("FONTSIZE",(0,1),(-1,-1),8),
                        ("GRID",(0,0),(-1,-1),0.25,MGRAY),
                        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
                        ("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),
                        ("ALIGN",(2,0),(-1,-1),"RIGHT"),
                    ]))
                    story.append(t_rev)
                    story.append(Spacer(1, 10))

                    # Cost-only table
                    if cost_only_results or unmatched_spend > 0:
                        story.append(sec("Cost-Only Providers (Nursing / Allied Health)"))
                        story.append(hr())
                        cost_rows = [["Provider","Discipline","Labor Cost","Note"]]
                        for cr in cost_only_results:
                            cost_rows.append([cr["provider"], cr["discipline"], fmt(cr["labor_cost"]), "Cost line only"])
                        if unmatched_spend > 0:
                            cost_rows.append(["Other / Unmatched","—",fmt(unmatched_spend),"Not matched to placed provider"])
                        t_cost = Table(cost_rows, colWidths=[1.6*inch,1.2*inch,1.0*inch,2.85*inch])
                        t_cost.setStyle(TableStyle([
                            ("BACKGROUND",(0,0),(-1,0),NAVY),("TEXTCOLOR",(0,0),(-1,0),WHITE),
                            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,0),8),
                            ("ROWBACKGROUNDS",(0,1),(-1,-1),[WHITE,LGRAY]),
                            ("FONTNAME",(0,1),(-1,-1),"Helvetica"),("FONTSIZE",(0,1),(-1,-1),8),
                            ("GRID",(0,0),(-1,-1),0.25,MGRAY),
                            ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
                            ("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),
                            ("ALIGN",(2,0),(2,-1),"RIGHT"),
                        ]))
                        story.append(t_cost)
                        story.append(Spacer(1, 10))

                    # Methodology
                    story.append(sec("Methodology"))
                    story.append(hr())
                    story.append(Paragraph(
                        "Labor cost: actual invoiced spend from the RMCHCS Spend Tracker for the selected period. "
                        "Estimated revenue: weekly volume × reimbursement rate × weeks in period. "
                        "Reimbursement rates sourced from DiagnOS benchmarks (HST Pathways 2024 ASC Report; "
                        "CMS RVU-based blended net collections; MGMA 2024; Medscape 2025). "
                        "Override rates with RMCHCS actual payer data for maximum accuracy. "
                        "Net contribution = estimated revenue minus total labor cost (including all disciplines). "
                        "Program ROI = net contribution ÷ true total cost.",
                        ps("m", fs=8, tc=MUTED, ld=11)))

                    buf2 = _io.BytesIO()
                    doc2 = SimpleDocTemplate(buf2, pagesize=letter,
                        leftMargin=0.65*inch, rightMargin=0.60*inch,
                        topMargin=0.88*inch, bottomMargin=0.68*inch)
                    doc2.build(story, onFirstPage=on_page, onLaterPages=on_page)
                    buf2.seek(0)
                    return buf2.read()

                # ── Build detailed PDF with charts ────────────────────────────────
                def build_roi_detailed_pdf(total_cost, total_rev, total_net, prog_roi,
                                            results, cost_only_results, unmatched_spend,
                                            start_date, end_date, period_weeks):
                    import io as _io
                    from reportlab.lib.pagesizes import letter
                    from reportlab.lib import colors
                    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                    from reportlab.lib.units import inch
                    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, Image as RLImage
                    from reportlab.lib.utils import ImageReader
                    from datetime import date as _date
                    import matplotlib
                    matplotlib.use("Agg")
                    import matplotlib.pyplot as plt
                    import matplotlib.patches as mpatches
                    import numpy as np

                    NAVY  = colors.HexColor("#283652")
                    RED   = colors.HexColor("#DD4F20")
                    WHITE = colors.white
                    LGRAY = colors.HexColor("#F1ECE9")
                    MGRAY = colors.HexColor("#B3B5B0")
                    GREEN = colors.HexColor("#065F46")
                    MUTED = colors.HexColor("#6B7280")
                    GREENBG = colors.HexColor("#ECFDF5")
                    W = 6.5 * inch

                    def fmt(v):  return f"${v:,.0f}"
                    def abbrev(v):
                        if abs(v) >= 1_000_000: return f"${v/1_000_000:.1f}M"
                        if abs(v) >= 1_000:     return f"${v/1_000:.0f}K"
                        return fmt(v)

                    def on_page(canvas, doc):
                        canvas.saveState()
                        pw, ph = letter
                        m = 0.75 * inch
                        canvas.setFillColor(NAVY)
                        canvas.rect(0, ph - 1.0*inch, pw, 1.0*inch, fill=1, stroke=0)
                        canvas.setFont("Helvetica-Bold", 11)
                        canvas.setFillColor(WHITE)
                        canvas.drawString(m, ph - 0.38*inch, "RMCHCS ITO Program")
                        canvas.setFont("Helvetica", 8)
                        canvas.setFillColor(colors.HexColor("#A8C4D4"))
                        canvas.drawString(m, ph - 0.58*inch,
                            f"Program ROI Report  |  {start_date.strftime('%b %d, %Y')} – {end_date.strftime('%b %d, %Y')}  |  Ingenovis Health")
                        canvas.drawRightString(pw - m, ph - 0.58*inch, f"Generated {_date.today().strftime('%B %d, %Y')}")
                        canvas.setStrokeColor(RED)
                        canvas.setLineWidth(1.5)
                        canvas.line(m, ph - 1.05*inch, pw - m, ph - 1.05*inch)
                        canvas.setStrokeColor(NAVY)
                        canvas.setLineWidth(0.5)
                        canvas.line(m, 0.62*inch, pw - m, 0.62*inch)
                        canvas.setFont("Helvetica", 7)
                        canvas.setFillColor(colors.HexColor("#777777"))
                        canvas.drawString(m, 0.44*inch, "Confidential  |  Ingenovis Health ITO  |  Revenue figures are estimated benchmarks")
                        canvas.drawRightString(pw - m, 0.44*inch, f"Page {doc.page}")
                        canvas.restoreState()

                    sty = getSampleStyleSheet()
                    def ps(name, fn="Helvetica", fs=9, ld=13, tc=None, **kw):
                        return ParagraphStyle(name, parent=sty["Normal"], fontName=fn,
                                              fontSize=fs, leading=ld,
                                              textColor=tc if tc is not None else colors.HexColor("#1A1A2E"), **kw)
                    def hr(c=NAVY, t=1.0): return HRFlowable(width="100%", thickness=t, color=c, spaceAfter=6, spaceBefore=4)

                    def h2(text):
                        _navy = colors.HexColor("#283652")
                        _lgray = colors.HexColor("#F1ECE9")
                        _h2style = ParagraphStyle("h2i", parent=sty["Normal"],
                            fontName="Helvetica-Bold", fontSize=10,
                            textColor=_navy, leading=14, leftIndent=8)
                        inner = Table([[Paragraph(text, _h2style)]], colWidths=[W])
                        inner.setStyle(TableStyle([
                            ("BACKGROUND",(0,0),(-1,-1),_lgray),
                            ("LINEBEFORE",(0,0),(0,-1),3,_navy),
                            ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
                            ("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),8),
                        ]))
                        return inner

                    def T(data, cw=None):
                        if cw is None: cw = [W*0.5, W*0.5]
                        t = Table(data, colWidths=cw)
                        t.setStyle(TableStyle([
                            ("BACKGROUND",(0,0),(-1,0),NAVY),("TEXTCOLOR",(0,0),(-1,0),WHITE),
                            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,0),9),
                            ("ROWBACKGROUNDS",(0,1),(-1,-1),[WHITE,LGRAY]),
                            ("FONTNAME",(0,1),(-1,-1),"Helvetica"),("FONTSIZE",(0,1),(-1,-1),9),
                            ("TEXTCOLOR",(0,1),(-1,-1),colors.HexColor("#222222")),
                            ("ALIGN",(1,0),(1,-1),"RIGHT"),
                            ("GRID",(0,0),(-1,-1),0.25,MGRAY),
                            ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
                            ("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8),
                        ]))
                        return t

                    def make_chart_img(width_in, height_in):
                        """Program-level cost vs revenue bar chart."""
                        fig, ax = plt.subplots(figsize=(width_in, height_in))
                        labels = ["Total Labor Cost", "Estimated Revenue"]
                        vals   = [total_cost, total_rev]
                        bar_colors = ["#ef4444", "#10b981"]
                        bars = ax.bar(labels, vals, color=bar_colors, width=0.45, edgecolor="white")
                        for bar, val in zip(bars, vals):
                            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + max(vals)*0.01,
                                    abbrev(val), ha="center", va="bottom", fontsize=9, fontweight="bold",
                                    color="#1a1a2e")
                        ax.set_facecolor("#f8fafc")
                        fig.patch.set_facecolor("#f8fafc")
                        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x,_: abbrev(x)))
                        ax.spines["top"].set_visible(False)
                        ax.spines["right"].set_visible(False)
                        ax.set_title("Program: Total Cost vs Estimated Revenue", fontsize=10,
                                     fontweight="bold", color="#283652", pad=8)
                        plt.tight_layout()
                        img_buf = _io.BytesIO()
                        plt.savefig(img_buf, format="png", dpi=150, bbox_inches="tight",
                                    facecolor="#f8fafc")
                        plt.close()
                        img_buf.seek(0)
                        return img_buf

                    def make_provider_chart(width_in, height_in):
                        """Per-provider cost vs revenue."""
                        if not results:
                            return None
                        pnames  = [r["Provider"] for r in results]
                        costs_v = [r["Labor Cost"] for r in results]
                        revs_v  = [r["Est. Revenue"] for r in results]
                        x = np.arange(len(pnames))
                        w = 0.35
                        fig, ax = plt.subplots(figsize=(width_in, height_in))
                        b1 = ax.bar(x - w/2, costs_v, w, label="Labor Cost", color="#ef4444", edgecolor="white")
                        b2 = ax.bar(x + w/2, revs_v,  w, label="Est. Revenue", color="#10b981", edgecolor="white")
                        ax.set_xticks(x)
                        ax.set_xticklabels(pnames, fontsize=8)
                        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x,_: abbrev(x)))
                        ax.set_facecolor("#f8fafc")
                        fig.patch.set_facecolor("#f8fafc")
                        ax.spines["top"].set_visible(False)
                        ax.spines["right"].set_visible(False)
                        ax.legend(fontsize=8)
                        ax.set_title("Revenue vs Labor Cost by MD/APP Provider", fontsize=10,
                                     fontweight="bold", color="#283652", pad=8)
                        plt.tight_layout()
                        img_buf = _io.BytesIO()
                        plt.savefig(img_buf, format="png", dpi=150, bbox_inches="tight", facecolor="#f8fafc")
                        plt.close()
                        img_buf.seek(0)
                        return img_buf

                    roi_c  = GREEN if prog_roi >= 0 else RED
                    roi_bg = GREENBG if prog_roi >= 0 else colors.HexColor("#FDEDEC")

                    story = []

                    # Title block
                    story.append(Paragraph("RMCHCS ITO Program", ps("t", fn="Helvetica-Bold", fs=20, tc=NAVY, ld=24, spaceBefore=2)))
                    story.append(Paragraph("Program ROI — Detailed Report",
                        ps("sub", fs=12, tc=colors.HexColor("#415B80"), ld=15, spaceAfter=3)))
                    story.append(Paragraph(
                        f"Period: {start_date.strftime('%B %d, %Y')} – {end_date.strftime('%B %d, %Y')}  ({period_weeks:.1f} weeks)  |  Ingenovis Health ITO  |  Generated {_date.today().strftime('%B %d, %Y')}",
                        ps("meta", fs=8, tc=MUTED, spaceAfter=8)))
                    story.append(HRFlowable(width="100%", thickness=2.0, color=RED, spaceAfter=12))

                    # Program summary table
                    story.append(h2("Program Summary"))
                    story.append(Spacer(1,6))
                    story.append(T([
                        ["Metric","Value"],
                        ["Analysis Period", f"{start_date.strftime('%b %d, %Y')} – {end_date.strftime('%b %d, %Y')} ({period_weeks:.1f} wks)"],
                        ["True Total Labor Cost", fmt(total_cost)],
                        ["Estimated Revenue Generated", fmt(total_rev)],
                        ["Net Contribution", fmt(total_net)],
                        ["Program ROI", f"{prog_roi:.1f}%"],
                        ["Revenue Providers", str(len(results))],
                    ], cw=[W*0.55, W*0.45]))
                    story.append(Spacer(1,8))

                    # ROI headline box
                    story.append(Table([[Paragraph(
                        f"<b>{prog_roi:.1f}% Program ROI</b> — The RMCHCS ITO program generated an estimated "
                        f"{abbrev(total_net)} net contribution on {abbrev(total_cost)} in total labor cost "
                        f"over the {period_weeks:.1f}-week analysis period.",
                        ps("roi", tc=roi_c, ld=13))]],
                        colWidths=[W],
                        style=TableStyle([("BACKGROUND",(0,0),(-1,-1),roi_bg),
                            ("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8),
                            ("LEFTPADDING",(0,0),(-1,-1),10),
                            ("LINEABOVE",(0,0),(-1,0),2.5,roi_c)])))
                    story.append(Spacer(1,14))

                    # Program chart
                    import tempfile, os as _os
                    story.append(h2("Program Cost vs Revenue"))
                    story.append(Spacer(1,6))
                    chart_buf = make_chart_img(6.0, 2.8)
                    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as _tf:
                        _tf.write(chart_buf.read())
                        _chart_path = _tf.name
                    story.append(RLImage(_chart_path, width=W, height=W*0.43))
                    story.append(Spacer(1,12))

                    # Provider chart
                    if results:
                        story.append(h2("Revenue vs Cost by Provider (MD / APP)"))
                        story.append(Spacer(1,6))
                        prov_chart = make_provider_chart(6.0, 2.8)
                        if prov_chart:
                            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as _tf2:
                                _tf2.write(prov_chart.read())
                                _prov_chart_path = _tf2.name
                            story.append(RLImage(_prov_chart_path, width=W, height=W*0.43))
                        story.append(Spacer(1,12))

                    # Revenue provider detail table
                    story.append(h2("Revenue Provider Detail"))
                    story.append(Spacer(1,6))
                    rev_rows = [["Provider","Specialty","Labor Cost","Est. Revenue","Net Contribution","ROI %","Units/Wk","Rate/Unit"]]
                    for r in results:
                        roi_str = f"{r['ROI %']:.1f}%" if isinstance(r["ROI %"], (int,float)) else str(r["ROI %"])
                        rev_rows.append([r["Provider"], r["Specialty"],
                                          fmt(r["Labor Cost"]), fmt(r["Est. Revenue"]),
                                          fmt(r["Net Contribution"]), roi_str,
                                          str(int(r["Units/Wk"])), fmt(r["Rate/Unit"])])
                    t_rev = Table(rev_rows, colWidths=[1.1*inch,1.15*inch,0.8*inch,0.8*inch,0.9*inch,0.55*inch,0.6*inch,0.6*inch])
                    t_rev.setStyle(TableStyle([
                        ("BACKGROUND",(0,0),(-1,0),NAVY),("TEXTCOLOR",(0,0),(-1,0),WHITE),
                        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,0),7.5),
                        ("ROWBACKGROUNDS",(0,1),(-1,-1),[WHITE,LGRAY]),
                        ("FONTNAME",(0,1),(-1,-1),"Helvetica"),("FONTSIZE",(0,1),(-1,-1),8),
                        ("GRID",(0,0),(-1,-1),0.25,MGRAY),
                        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
                        ("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(0,0),(-1,-1),5),
                        ("ALIGN",(2,0),(-1,-1),"RIGHT"),
                    ]))
                    story.append(t_rev)
                    story.append(Spacer(1,10))

                    # Cost-only table
                    if cost_only_results or unmatched_spend > 0:
                        story.append(h2("Cost-Only Providers (Nursing / Allied Health)"))
                        story.append(Spacer(1,6))
                        cost_rows = [["Provider","Discipline","Labor Cost","Note"]]
                        for cr in cost_only_results:
                            cost_rows.append([cr["provider"], cr["discipline"], fmt(cr["labor_cost"]), "Cost line only — no revenue attribution"])
                        if unmatched_spend > 0:
                            cost_rows.append(["Other / Unmatched","—",fmt(unmatched_spend),"Logged spend not matched to placed provider"])
                        t_cost = Table(cost_rows, colWidths=[1.5*inch,1.1*inch,0.9*inch,3.0*inch])
                        t_cost.setStyle(TableStyle([
                            ("BACKGROUND",(0,0),(-1,0),NAVY),("TEXTCOLOR",(0,0),(-1,0),WHITE),
                            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,0),8),
                            ("ROWBACKGROUNDS",(0,1),(-1,-1),[WHITE,LGRAY]),
                            ("FONTNAME",(0,1),(-1,-1),"Helvetica"),("FONTSIZE",(0,1),(-1,-1),8),
                            ("GRID",(0,0),(-1,-1),0.25,MGRAY),
                            ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
                            ("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),
                            ("ALIGN",(2,0),(2,-1),"RIGHT"),
                        ]))
                        story.append(t_cost)
                        story.append(Spacer(1,10))

                    # Methodology
                    story.append(h2("Methodology & Data Sources"))
                    story.append(Spacer(1,6))
                    story.append(Paragraph(
                        "<b>Labor cost:</b> Actual invoiced spend from the RMCHCS Spend Tracker for the selected period. "
                        "<b>Estimated revenue:</b> Weekly procedure/encounter volume × reimbursement rate × weeks in period. "
                        "<b>Reimbursement rates:</b> DiagnOS benchmarks sourced from HST Pathways 2024 ASC State of the Industry Report (590 surgery centers, ~3M patient visits); "
                        "CMS RVU-based blended net collections; MGMA 2024 Provider Compensation Data Report; Medscape 2025 Revenue & Compensation Report. "
                        "Override with RMCHCS actual contracted payer rates for maximum accuracy. "
                        "<b>Net contribution</b> = estimated revenue minus true total labor cost (all disciplines). "
                        "<b>Program ROI</b> = net contribution ÷ true total cost.",
                        ps("m", fs=8, tc=MUTED, ld=11)))

                    buf2 = _io.BytesIO()
                    doc2 = SimpleDocTemplate(buf2, pagesize=letter,
                        leftMargin=0.75*inch, rightMargin=0.75*inch,
                        topMargin=1.2*inch, bottomMargin=0.9*inch)
                    doc2.build(story, onFirstPage=on_page, onLaterPages=on_page)
                    # Clean up temp image files
                    try:
                        if '_chart_path' in dir(): _os.unlink(_chart_path)
                        if '_prov_chart_path' in dir(): _os.unlink(_prov_chart_path)
                    except Exception:
                        pass
                    buf2.seek(0)
                    return buf2.read()

                # ── Download buttons ──────────────────────────────────────────────
                with ex1:
                    exec_pdf = build_roi_exec_brief(
                        true_total_cost, total_rev, total_net, prog_roi,
                        results, cost_only_results, unmatched_spend,
                        start_date, end_date, period_weeks)
                    st.download_button(
                        "📄 Executive Brief",
                        data=exec_pdf,
                        file_name=f"RMCHCS_ITO_Executive_Brief_{date.today().strftime('%Y%m%d')}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                        type="primary"
                    )

                with ex2:
                    detailed_pdf = build_roi_detailed_pdf(
                        true_total_cost, total_rev, total_net, prog_roi,
                        results, cost_only_results, unmatched_spend,
                        start_date, end_date, period_weeks)
                    st.download_button(
                        "📊 Detailed Report",
                        data=detailed_pdf,
                        file_name=f"RMCHCS_ITO_Detailed_Report_{date.today().strftime('%Y%m%d')}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                        type="secondary"
                    )

                st.markdown(
                    '<div style="background:#f1f5f9;border-radius:8px;padding:12px 16px;'
                    'font-size:0.75rem;color:#64748b;margin-top:8px;">'
                    '<strong>Methodology:</strong> Labor cost = actual invoiced spend from the Spend Tracker '
                    'for the selected period. Estimated revenue = weekly volume × reimbursement rate × weeks '
                    'in period. Reimbursement rates are CMS-sourced benchmarks from DiagnOS '
                    '(HST Pathways 2024, CMS RVU-based blended net collections, MGMA 2024, Medscape 2025). '
                    'Override rates with RMCHCS actual payer data for maximum accuracy. '
                    'Net contribution = estimated revenue minus labor cost. '
                    'Program ROI = net contribution ÷ labor cost.'
                    '</div>',
                    unsafe_allow_html=True
                )


elif page == "─────────────":
    st.info("Select a section from the sidebar.")
